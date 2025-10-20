# Standard library imports
from datetime import datetime, date
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Avg, Case, Count, IntegerField, Q, Value, When
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.core.paginator import Paginator

# Third-party imports
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Local project imports
from management_project.forms import StakeholderForm
from management_project.models import Stakeholder, models

#
@login_required
def stakeholder_list(request):
    query = request.GET.get('search', '').strip()
    selected_type = request.GET.get('stakeholder_type', '').strip()
    page_number = request.GET.get('page', 1)

    # Base queryset: all stakeholders for this organization
    stakeholders = Stakeholder.objects.filter(
        organization_name=request.user.organization_name
    )

    # Apply exact filter by stakeholder type (internal/external)
    if selected_type:
        stakeholders = stakeholders.filter(stakeholder_type=selected_type)

    # Apply search filter across multiple fields
    if query:
        search_filter = (
            Q(stakeholder_name__icontains=query) |
            Q(role__icontains=query) |
            Q(department__icontains=query) |  # FK field now properly searched
            Q(description__icontains=query) |
            Q(contact_info__icontains=query)
        )
        stakeholders = stakeholders.filter(search_filter)

    # Ordering
    stakeholders = stakeholders.order_by('stakeholder_type', 'priority', 'stakeholder_name')

    # Provide stakeholder types for dropdown filter
    stakeholder_types = [choice[0] for choice in Stakeholder.STAKEHOLDER_TYPE_CHOICES]

    # Pagination
    paginator = Paginator(stakeholders, 10)
    page_obj = paginator.get_page(page_number)

    return render(request, 'stakeholder_list/list.html', {
        'stakeholders': page_obj,
        'page_obj': page_obj,
        'search_query': query,
        'stakeholder_types': stakeholder_types,
        'selected_type': selected_type,
    })


@login_required
def create_stakeholder(request):
    """
    Create a new stakeholder entry for the current user's organization.
    Redirects back to the child form if 'next' is provided,
    optionally preselecting the newly created stakeholder.
    """
    next_url = request.GET.get("next") or request.POST.get("next")  # URL to return to child

    if request.method == "POST":
        form = StakeholderForm(request.POST, request=request)
        if form.is_valid():
            stakeholder = form.save(commit=False)
            stakeholder.organization_name = request.user.organization_name
            stakeholder.save()
            form.save_m2m()
            messages.success(request, "Stakeholder created successfully!")

            # Redirect back to child form with new stakeholder preselected
            if next_url:
                separator = '&' if '?' in next_url else '?'
                return redirect(f"{next_url}{separator}stakeholder={stakeholder.pk}")

            return redirect("stakeholder_list")  # fallback to parent list

    else:
        form = StakeholderForm(request=request)

    return render(request, "stakeholder_list/form.html", {
        "form": form,
        "next": next_url,
    })



@login_required
def update_stakeholder(request, pk):
    stakeholder = get_object_or_404(Stakeholder, pk=pk, organization_name=request.user.organization_name)

    if request.method == 'POST':
        form = StakeholderForm(request.POST, instance=stakeholder, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, "Stakeholder updated successfully!")
            return redirect('stakeholder_list')
    else:
        form = StakeholderForm(instance=stakeholder, request=request)

    context = {
        'form': form,
        'form_title': 'Update Stakeholder',
        'submit_button_text': 'Update Stakeholder',
        'back_url': request.GET.get('next', reverse('stakeholder_list')),
        'edit_mode': True,
        'editing_stakeholder': stakeholder,
    }
    return render(request, 'stakeholder_list/form.html', context)


@login_required
def delete_stakeholder(request, pk):
    stakeholder = get_object_or_404(Stakeholder, pk=pk, organization_name=request.user.organization_name)

    if request.method == 'POST':
        stakeholder.delete()
        messages.success(request, "Stakeholder deleted successfully!")
        return redirect('stakeholder_list')

    return render(request, 'stakeholder_list/delete_confirm.html', {'stakeholder': stakeholder})



@login_required
def export_stakeholders_to_excel(request):
    query = request.GET.get('search', '').strip()
    selected_type = request.GET.get('stakeholder_type', '').strip()

    stakeholders = Stakeholder.objects.all()

    # Filter by stakeholder type
    if selected_type:
        stakeholders = stakeholders.filter(stakeholder_type=selected_type)

    # Apply search filter
    if query:
        search_filter = (
            Q(stakeholder_name__icontains=query) |
            Q(role__icontains=query) |
            Q(department__icontains=query) |
            Q(description__icontains=query) |
            Q(notes__icontains=query) |
            Q(contact_info__icontains=query)
        )
        stakeholders = stakeholders.filter(search_filter)

    # Helper to get field values
    def get_field_value(instance, field_name):
        value = getattr(instance, field_name, "")
        # ManyToManyField: show only objective
        if field_name == "aligned_objectives":
            value = ", ".join(str(obj.objective) for obj in value.all() if getattr(obj, 'objective', None))
        elif isinstance(value, models.Model):
            value = str(value)
        elif isinstance(value, (list, tuple)):
            value = ", ".join(map(str, value))
        elif isinstance(value, (datetime, date)):
            value = value.strftime("%d %B, %Y")  # Format like 12 July, 2023
        elif isinstance(value, bool):
            value = "Yes" if value else "No"
        elif value is None:
            value = ""
        return value

    # Fields in the same order as the model (without organization_name & slug)
    field_names = [
        'stakeholder_code', 'stakeholder_name', 'stakeholder_type', 'stakeholder_category',
        'role', 'primary_role',
        'impact_level', 'interest_level', 'influence_score', 'risk_level', 'contribution_score',
        'priority', 'satisfaction_level', 'engagement_strategy',
        'email', 'phone', 'contact_info', 'department', 'location',
        'description', 'notes',
        'relationship_status', 'last_engagement_date', 'next_engagement_date',
        'aligned_objectives',
        'engagement_priority_score', 'is_key_stakeholder', 'requires_attention',
        'created_at', 'updated_at'
    ]

    headers = [field.replace('_', ' ').title() for field in field_names]

    # Workbook setup
    wb = Workbook()
    ws = wb.active
    ws.title = "Stakeholders"

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="Stakeholder List")
    title_cell.font = Font(size=14, bold=True, color="FFFFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color="FF305496", end_color="FF305496", fill_type="solid")

    # Headers
    ws.append(headers)
    for cell in ws[2]:
        cell.fill = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for s in stakeholders:
        row = [get_field_value(s, field) for field in field_names]
        ws.append(row)

    # Adjust column widths
    for i, column_cells in enumerate(ws.columns, 1):
        max_length = max((len(str(cell.value)) for cell in column_cells if cell.value), default=0) + 2
        ws.column_dimensions[get_column_letter(i)].width = max(max_length, 12)

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=stakeholders.xlsx'
    wb.save(response)
    return response


#
# # Model imports
# # views.py
# # management_project/views/stakeholder.py
# @login_required
# def stakeholder_graph_view(request):
#     qs = Stakeholder.objects.filter(organization_name=request.user.organization_name)
#
#     # Enhanced color palette with better contrast
#     color_palette = [
#         "#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd",
#         "#8c564b", "#e377c2", "#7f7f7f", "#bcbd22", "#17becf",
#         "#aec7e8", "#ffbb78", "#98df8a", "#ff9896", "#c5b0d5"
#     ]
#
#     # ------------------ STAKEHOLDER CATEGORY DISTRIBUTION ------------------
#     category_data = qs.values('stakeholder_category').annotate(
#         count=Count('id'),
#         avg_engagement=Avg('engagement_priority_score'),
#         high_risk_count=Count('id', filter=Q(risk_level__in=['high', 'very_high']))
#     ).order_by('-count')
#
#     category_labels = [cat['stakeholder_category'].replace('_', ' ').title() for cat in category_data]
#     category_counts = [cat['count'] for cat in category_data]
#     category_avg_engagement = [cat['avg_engagement'] or 0 for cat in category_data]
#     category_high_risk = [cat['high_risk_count'] for cat in category_data]
#
#     fig_category = go.Figure()
#     fig_category.add_trace(go.Bar(
#         name='Stakeholder Count',
#         x=category_labels,
#         y=category_counts,
#         marker_color=color_palette[0],
#         text=category_counts,
#         textposition='auto',
#     ))
#     fig_category.add_trace(go.Scatter(
#         name='Avg Engagement Score',
#         x=category_labels,
#         y=category_avg_engagement,
#         mode='lines+markers',
#         line=dict(color=color_palette[1], width=3),
#         yaxis='y2'
#     ))
#     fig_category.update_layout(
#         title='Stakeholder Category Analysis<br><sub>Bars show count, line shows average engagement score</sub>',
#         xaxis_title='Stakeholder Category',
#         yaxis=dict(title='Count', side='left'),
#         yaxis2=dict(title='Engagement Score', side='right', overlaying='y'),
#         template='plotly_white',
#         height=500,
#         showlegend=True
#     )
#
#     # ------------------ ROLE DISTRIBUTION (Top 10) ------------------
#     role_data = {}
#     for stakeholder in qs:
#         roles = stakeholder.role or ['unassigned']
#         for role in roles:
#             label = role.replace('_', ' ').title() if role != 'other' else 'Other'
#             role_data[label] = role_data.get(label, 0) + 1
#
#     # Get top 10 roles by count
#     sorted_roles = sorted(role_data.items(), key=lambda x: x[1], reverse=True)[:10]
#     role_labels = [role[0] for role in sorted_roles]
#     role_values = [role[1] for role in sorted_roles]
#
#     fig_role = go.Figure(data=[go.Bar(
#         x=role_values,
#         y=role_labels,
#         orientation='h',
#         marker_color=color_palette[:len(role_labels)],
#         text=role_values,
#         textposition='auto',
#         hovertemplate='<b>%{y}</b><br>Count: %{x}<extra></extra>'
#     )])
#     fig_role.update_layout(
#         title='Top 10 Stakeholder Roles',
#         xaxis_title='Count',
#         yaxis_title='Role',
#         template='plotly_white',
#         height=500
#     )
#
#     # ------------------ STAKEHOLDER TYPE ANALYSIS ------------------
#     type_analysis = qs.values('stakeholder_type').annotate(
#         count=Count('id'),
#         avg_engagement=Avg('engagement_priority_score'),
#         avg_satisfaction=Avg(Case(
#             When(satisfaction_level='very_low', then=1),
#             When(satisfaction_level='low', then=2),
#             When(satisfaction_level='medium', then=3),
#             When(satisfaction_level='high', then=4),
#             When(satisfaction_level='very_high', then=5),
#             default=3,
#             output_field=IntegerField()
#         )),
#         key_stakeholder_count=Count('id', filter=Q(is_key_stakeholder=True))
#     )
#
#     type_labels = [t['stakeholder_type'].replace('_', ' ').title() for t in type_analysis]
#     type_counts = [t['count'] for t in type_analysis]
#     type_engagement = [t['avg_engagement'] or 0 for t in type_analysis]
#     type_satisfaction = [t['avg_satisfaction'] or 0 for t in type_analysis]
#     type_key_stakeholders = [t['key_stakeholder_count'] for t in type_analysis]
#
#     fig_type = go.Figure()
#     fig_type.add_trace(go.Bar(
#         name='Total Stakeholders',
#         x=type_labels,
#         y=type_counts,
#         marker_color=color_palette[0],
#         text=type_counts,
#         textposition='auto'
#     ))
#     fig_type.add_trace(go.Bar(
#         name='Key Stakeholders',
#         x=type_labels,
#         y=type_key_stakeholders,
#         marker_color=color_palette[1],
#         text=type_key_stakeholders,
#         textposition='auto'
#     ))
#     fig_type.update_layout(
#         title='Stakeholder Type Analysis<br><sub>Showing total and key stakeholders by type</sub>',
#         xaxis_title='Stakeholder Type',
#         yaxis_title='Count',
#         template='plotly_white',
#         barmode='group',
#         height=500
#     )
#
#     # ------------------ RISK vs IMPACT MATRIX ------------------
#     risk_impact_data = {}
#     for stakeholder in qs:
#         risk = stakeholder.risk_level or 'medium'
#         impact = stakeholder.impact_level or 'medium'
#         key = (risk, impact)
#         risk_impact_data[key] = risk_impact_data.get(key, 0) + 1
#
#     # Prepare data for heatmap
#     risk_levels = ['very_low', 'low', 'medium', 'high', 'very_high']
#     impact_levels = ['very_low', 'low', 'medium', 'high', 'very_high']
#
#     heatmap_data = []
#     for risk in risk_levels:
#         row = []
#         for impact in impact_levels:
#             row.append(risk_impact_data.get((risk, impact), 0))
#         heatmap_data.append(row)
#
#     fig_risk_impact = go.Figure(data=go.Heatmap(
#         z=heatmap_data,
#         x=[level.replace('_', ' ').title() for level in impact_levels],
#         y=[level.replace('_', ' ').title() for level in risk_levels],
#         colorscale='RdYlGn_r',  # Red for high risk/impact, Green for low
#         hoverongaps=False,
#         hovertemplate='Risk: %{y}<br>Impact: %{x}<br>Count: %{z}<extra></extra>',
#         text=heatmap_data,
#         texttemplate="%{z}",
#         textfont={"size": 12}
#     ))
#     fig_risk_impact.update_layout(
#         title='Risk vs Impact Matrix<br><sub>Number of stakeholders in each risk-impact category</sub>',
#         xaxis_title='Impact Level',
#         yaxis_title='Risk Level',
#         template='plotly_white',
#         height=500
#     )
#
#     # ------------------ ENGAGEMENT STRATEGY EFFECTIVENESS ------------------
#     engagement_effectiveness = {}
#     for stakeholder in qs:
#         satisfaction_value = {
#             'very_low': 1, 'low': 2, 'medium': 3, 'high': 4, 'very_high': 5
#         }.get(stakeholder.satisfaction_level, 3)
#
#         strategies = stakeholder.engagement_strategy or ['none']
#         for strategy in strategies:
#             label = strategy.replace('_', ' ').title()
#             if label not in engagement_effectiveness:
#                 engagement_effectiveness[label] = {
#                     'count': 0,
#                     'total_satisfaction': 0,
#                     'avg_engagement_score': 0,
#                     'engagement_scores': []
#                 }
#             engagement_effectiveness[label]['count'] += 1
#             engagement_effectiveness[label]['total_satisfaction'] += satisfaction_value
#             engagement_effectiveness[label]['engagement_scores'].append(float(stakeholder.engagement_priority_score))
#
#     # Calculate averages
#     for strategy in engagement_effectiveness.values():
#         strategy['avg_satisfaction'] = strategy['total_satisfaction'] / strategy['count']
#         strategy['avg_engagement_score'] = sum(strategy['engagement_scores']) / len(strategy['engagement_scores'])
#
#     engagement_labels = list(engagement_effectiveness.keys())
#     engagement_counts = [data['count'] for data in engagement_effectiveness.values()]
#     engagement_satisfaction = [data['avg_satisfaction'] for data in engagement_effectiveness.values()]
#     engagement_scores = [data['avg_engagement_score'] for data in engagement_effectiveness.values()]
#
#     fig_engagement = go.Figure()
#     fig_engagement.add_trace(go.Bar(
#         name='Stakeholder Count',
#         x=engagement_labels,
#         y=engagement_counts,
#         marker_color=color_palette[0],
#         text=engagement_counts,
#         textposition='auto',
#     ))
#     fig_engagement.add_trace(go.Scatter(
#         name='Avg Satisfaction',
#         x=engagement_labels,
#         y=engagement_satisfaction,
#         mode='lines+markers',
#         line=dict(color=color_palette[1], width=3),
#         yaxis='y2'
#     ))
#     fig_engagement.update_layout(
#         title='Engagement Strategy Effectiveness<br><sub>Bars show usage count, line shows average satisfaction</sub>',
#         xaxis_title='Engagement Strategy',
#         yaxis=dict(title='Count', side='left'),
#         yaxis2=dict(title='Average Satisfaction (1-5)', side='right', overlaying='y', range=[1, 5]),
#         template='plotly_white',
#         height=500
#     )
#
#     # ------------------ PRIORITY ANALYSIS ------------------
#     priority_analysis = qs.values('priority').annotate(
#         count=Count('id'),
#         avg_engagement=Avg('engagement_priority_score'),
#         requires_attention_count=Count('id', filter=Q(requires_attention=True)),
#         key_stakeholder_count=Count('id', filter=Q(is_key_stakeholder=True))
#     ).order_by('priority')
#
#     priority_labels = [p['priority'].replace('_', ' ').title() for p in priority_analysis]
#     priority_counts = [p['count'] for p in priority_analysis]
#     priority_attention = [p['requires_attention_count'] for p in priority_analysis]
#     priority_key = [p['key_stakeholder_count'] for p in priority_analysis]
#
#     fig_priority = go.Figure()
#     fig_priority.add_trace(go.Bar(
#         name='Total Stakeholders',
#         x=priority_labels,
#         y=priority_counts,
#         marker_color=color_palette[0],
#         text=priority_counts,
#         textposition='auto'
#     ))
#     fig_priority.add_trace(go.Bar(
#         name='Require Attention',
#         x=priority_labels,
#         y=priority_attention,
#         marker_color=color_palette[1],
#         text=priority_attention,
#         textposition='auto'
#     ))
#     fig_priority.add_trace(go.Bar(
#         name='Key Stakeholders',
#         x=priority_labels,
#         y=priority_key,
#         marker_color=color_palette[2],
#         text=priority_key,
#         textposition='auto'
#     ))
#     fig_priority.update_layout(
#         title='Stakeholder Priority Analysis<br><sub>Breakdown by priority level with attention and key stakeholder flags</sub>',
#         xaxis_title='Priority Level',
#         yaxis_title='Count',
#         template='plotly_white',
#         barmode='group',
#         height=500
#     )
#
#     # ------------------ RELATIONSHIP STATUS OVERVIEW ------------------
#     # Calculate overdue engagements manually since it's a property
#     from datetime import date, timedelta
#     ninety_days_ago = date.today() - timedelta(days=90)
#
#     relationship_data = qs.values('relationship_status').annotate(
#         count=Count('id'),
#         avg_satisfaction=Avg(Case(
#             When(satisfaction_level='very_low', then=1),
#             When(satisfaction_level='low', then=2),
#             When(satisfaction_level='medium', then=3),
#             When(satisfaction_level='high', then=4),
#             When(satisfaction_level='very_high', then=5),
#             default=3,
#             output_field=IntegerField()
#         ))
#     ).order_by('relationship_status')
#
#     # Calculate overdue engagements for each relationship status
#     relationship_overdue = {}
#     for status in relationship_data:
#         status_qs = qs.filter(relationship_status=status['relationship_status'])
#         overdue_count = 0
#         for stakeholder in status_qs:
#             if (stakeholder.last_engagement_date and
#                     stakeholder.last_engagement_date < ninety_days_ago):
#                 overdue_count += 1
#         relationship_overdue[status['relationship_status']] = overdue_count
#
#     relationship_labels = [r['relationship_status'].title() for r in relationship_data]
#     relationship_counts = [r['count'] for r in relationship_data]
#     relationship_satisfaction = [r['avg_satisfaction'] or 0 for r in relationship_data]
#     relationship_overdue_counts = [relationship_overdue.get(status, 0) for status in
#                                    [r['relationship_status'] for r in relationship_data]]
#
#     fig_relationship = go.Figure()
#     fig_relationship.add_trace(go.Bar(
#         name='Total Stakeholders',
#         x=relationship_labels,
#         y=relationship_counts,
#         marker_color=color_palette[0],
#         text=relationship_counts,
#         textposition='auto'
#     ))
#     fig_relationship.add_trace(go.Bar(
#         name='Overdue Engagement',
#         x=relationship_labels,
#         y=relationship_overdue_counts,
#         marker_color=color_palette[1],
#         text=relationship_overdue_counts,
#         textposition='auto'
#     ))
#     fig_relationship.update_layout(
#         title='Relationship Status Overview<br><sub>Total stakeholders and overdue engagements by relationship status</sub>',
#         xaxis_title='Relationship Status',
#         yaxis_title='Count',
#         template='plotly_white',
#         barmode='group',
#         height=500
#     )
#
#     # ------------------ ENHANCED SUMMARY DATA ------------------
#     total_stakeholders = qs.count()
#
#     # Calculate percentages and averages with None handling
#     internal_count = qs.filter(stakeholder_type='internal').count()
#     external_count = qs.filter(stakeholder_type='external').count()
#     interface_count = qs.filter(stakeholder_type='interface').count()
#
#     # Risk analysis
#     high_risk_count = qs.filter(risk_level__in=['high', 'very_high']).count()
#     medium_risk_count = qs.filter(risk_level='medium').count()
#     low_risk_count = qs.filter(risk_level__in=['low', 'very_low']).count()
#
#     # Priority analysis
#     high_priority_count = qs.filter(priority__in=['high', 'very_high']).count()
#     key_stakeholder_count = qs.filter(is_key_stakeholder=True).count()
#     attention_required_count = qs.filter(requires_attention=True).count()
#
#     # Engagement analysis - calculate overdue manually
#     overdue_engagement_count = 0
#     for stakeholder in qs:
#         if (stakeholder.last_engagement_date and
#                 stakeholder.last_engagement_date < ninety_days_ago):
#             overdue_engagement_count += 1
#
#     avg_engagement_score = qs.aggregate(avg=Avg('engagement_priority_score'))['avg'] or 0
#
#     # Satisfaction analysis
#     high_satisfaction_count = qs.filter(satisfaction_level__in=['high', 'very_high']).count()
#     low_satisfaction_count = qs.filter(satisfaction_level__in=['low', 'very_low']).count()
#
#     summary_data = {
#         'total_stakeholders': total_stakeholders,
#         'internal_count': internal_count,
#         'external_count': external_count,
#         'interface_count': interface_count,
#         'internal_percentage': round((internal_count / total_stakeholders * 100) if total_stakeholders > 0 else 0, 1),
#         'external_percentage': round((external_count / total_stakeholders * 100) if total_stakeholders > 0 else 0, 1),
#         'high_risk_count': high_risk_count,
#         'high_risk_percentage': round((high_risk_count / total_stakeholders * 100) if total_stakeholders > 0 else 0, 1),
#         'medium_risk_percentage': round((medium_risk_count / total_stakeholders * 100) if total_stakeholders > 0 else 0,
#                                         1),
#         'low_risk_percentage': round((low_risk_count / total_stakeholders * 100) if total_stakeholders > 0 else 0, 1),
#         'high_priority_count': high_priority_count,
#         'key_stakeholder_count': key_stakeholder_count,
#         'attention_required_count': attention_required_count,
#         'attention_percentage': round(
#             (attention_required_count / total_stakeholders * 100) if total_stakeholders > 0 else 0, 1),
#         'overdue_engagement_count': overdue_engagement_count,
#         'overdue_percentage': round(
#             (overdue_engagement_count / total_stakeholders * 100) if total_stakeholders > 0 else 0, 1),
#         'avg_engagement_score': round(avg_engagement_score, 1),
#         'high_satisfaction_count': high_satisfaction_count,
#         'high_satisfaction_percentage': round(
#             (high_satisfaction_count / total_stakeholders * 100) if total_stakeholders > 0 else 0, 1),
#         'low_satisfaction_count': low_satisfaction_count,
#         'low_satisfaction_percentage': round(
#             (low_satisfaction_count / total_stakeholders * 100) if total_stakeholders > 0 else 0, 1),
#     }
#
#     return render(request, 'stakeholder_list/graph.html', {
#         'plot_html_category': fig_category.to_html(full_html=False),
#         'plot_html_role': fig_role.to_html(full_html=False),
#         'plot_html_type': fig_type.to_html(full_html=False),
#         'plot_html_risk_impact': fig_risk_impact.to_html(full_html=False),
#         'plot_html_engagement': fig_engagement.to_html(full_html=False),
#         'plot_html_priority': fig_priority.to_html(full_html=False),
#         'plot_html_relationship': fig_relationship.to_html(full_html=False),
#         'summary_data': summary_data,
#     })


@login_required
def stakeholder_graph_view(request):
    qs = Stakeholder.objects.filter(organization_name=request.user.organization_name)

    # Enhanced color palette with better contrast
    color_palette = [
        "#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd",
        "#8c564b", "#e377c2", "#7f7f7f", "#bcbd22", "#17becf",
        "#aec7e8", "#ffbb78", "#98df8a", "#ff9896", "#c5b0d5"
    ]

    # ------------------ STAKEHOLDER TYPE COUNT BOX ------------------
    type_counts = qs.values('stakeholder_type').annotate(count=Count('id'))
    type_data = {item['stakeholder_type']: item['count'] for item in type_counts}

    total_stakeholders = qs.count()
    internal_count = type_data.get('internal', 0)
    external_count = type_data.get('external', 0)
    interface_count = type_data.get('interface', 0)

    # Create stakeholder type count box
    fig_type_count = go.Figure()

    # Add indicators for each stakeholder type
    fig_type_count.add_trace(go.Indicator(
        mode="number",
        value=internal_count,
        number={'font': {'size': 28, 'color': color_palette[0]}},
        title={'text': "INTERNAL<br><span style='font-size:0.8em;color:gray'>Stakeholders</span>",
               'font': {'size': 14}},
        domain={'row': 0, 'column': 0}
    ))

    fig_type_count.add_trace(go.Indicator(
        mode="number",
        value=external_count,
        number={'font': {'size': 28, 'color': color_palette[1]}},
        title={'text': "EXTERNAL<br><span style='font-size:0.8em;color:gray'>Stakeholders</span>",
               'font': {'size': 14}},
        domain={'row': 0, 'column': 1}
    ))

    fig_type_count.add_trace(go.Indicator(
        mode="number",
        value=interface_count,
        number={'font': {'size': 28, 'color': color_palette[2]}},
        title={'text': "INTERFACE<br><span style='font-size:0.8em;color:gray'>Stakeholders</span>",
               'font': {'size': 14}},
        domain={'row': 0, 'column': 2}
    ))

    fig_type_count.add_trace(go.Indicator(
        mode="number",
        value=total_stakeholders,
        number={'font': {'size': 32, 'color': color_palette[4]}},
        title={'text': "TOTAL<br><span style='font-size:0.8em;color:gray'>Stakeholders</span>",
               'font': {'size': 16}},
        domain={'row': 1, 'column': 1}
    ))

    fig_type_count.update_layout(
        grid={'rows': 2, 'columns': 3, 'pattern': "independent"},
        template='plotly_white',
        height=300,
        margin=dict(l=20, r=20, t=50, b=20),
        title={
            'text': 'Stakeholder Type Distribution',
            'x': 0.5,
            'xanchor': 'center',
            'font': {'size': 20}
        }
    )

    # ------------------ STAKEHOLDER CATEGORY DISTRIBUTION ------------------
    category_data = qs.values('stakeholder_category').annotate(
        count=Count('id'),
        avg_engagement=Avg('engagement_priority_score'),
        high_risk_count=Count('id', filter=Q(risk_level__in=['high', 'very_high']))
    ).order_by('-count')

    category_labels = [cat['stakeholder_category'].replace('_', ' ').title() for cat in category_data]
    category_counts = [cat['count'] for cat in category_data]
    category_avg_engagement = [cat['avg_engagement'] or 0 for cat in category_data]
    category_high_risk = [cat['high_risk_count'] for cat in category_data]

    fig_category = go.Figure()
    fig_category.add_trace(go.Bar(
        name='Stakeholder Count',
        x=category_labels,
        y=category_counts,
        marker_color=color_palette[0],
        text=category_counts,
        textposition='auto',
    ))
    fig_category.add_trace(go.Scatter(
        name='Avg Engagement Score',
        x=category_labels,
        y=category_avg_engagement,
        mode='lines+markers',
        line=dict(color=color_palette[1], width=3),
        yaxis='y2'
    ))
    fig_category.update_layout(
        title='Stakeholder Category Analysis<br><sub>Bars show count, line shows average engagement score</sub>',
        xaxis_title='Stakeholder Category',
        yaxis=dict(title='Count', side='left'),
        yaxis2=dict(title='Engagement Score', side='right', overlaying='y'),
        template='plotly_white',
        height=500,
        showlegend=True
    )

    # ------------------ ROLE DISTRIBUTION (Top 10) ------------------
    role_data = {}
    for stakeholder in qs:
        roles = stakeholder.role or ['unassigned']
        for role in roles:
            label = role.replace('_', ' ').title() if role != 'other' else 'Other'
            role_data[label] = role_data.get(label, 0) + 1

    # Get top 10 roles by count
    sorted_roles = sorted(role_data.items(), key=lambda x: x[1], reverse=True)[:10]
    role_labels = [role[0] for role in sorted_roles]
    role_values = [role[1] for role in sorted_roles]

    fig_role = go.Figure(data=[go.Bar(
        x=role_values,
        y=role_labels,
        orientation='h',
        marker_color=color_palette[:len(role_labels)],
        text=role_values,
        textposition='auto',
        hovertemplate='<b>%{y}</b><br>Count: %{x}<extra></extra>'
    )])
    fig_role.update_layout(
        title='Top 10 Stakeholder Roles',
        xaxis_title='Count',
        yaxis_title='Role',
        template='plotly_white',
        height=500
    )

    # ------------------ STAKEHOLDER TYPE ANALYSIS ------------------
    type_analysis = qs.values('stakeholder_type').annotate(
        count=Count('id'),
        avg_engagement=Avg('engagement_priority_score'),
        avg_satisfaction=Avg(Case(
            When(satisfaction_level='very_low', then=1),
            When(satisfaction_level='low', then=2),
            When(satisfaction_level='medium', then=3),
            When(satisfaction_level='high', then=4),
            When(satisfaction_level='very_high', then=5),
            default=3,
            output_field=IntegerField()
        )),
        key_stakeholder_count=Count('id', filter=Q(is_key_stakeholder=True))
    )

    type_labels = [t['stakeholder_type'].replace('_', ' ').title() for t in type_analysis]
    type_counts = [t['count'] for t in type_analysis]
    type_engagement = [t['avg_engagement'] or 0 for t in type_analysis]
    type_satisfaction = [t['avg_satisfaction'] or 0 for t in type_analysis]
    type_key_stakeholders = [t['key_stakeholder_count'] for t in type_analysis]

    fig_type = go.Figure()
    fig_type.add_trace(go.Bar(
        name='Total Stakeholders',
        x=type_labels,
        y=type_counts,
        marker_color=color_palette[0],
        text=type_counts,
        textposition='auto'
    ))
    fig_type.add_trace(go.Bar(
        name='Key Stakeholders',
        x=type_labels,
        y=type_key_stakeholders,
        marker_color=color_palette[1],
        text=type_key_stakeholders,
        textposition='auto'
    ))
    fig_type.update_layout(
        title='Stakeholder Type Analysis<br><sub>Showing total and key stakeholders by type</sub>',
        xaxis_title='Stakeholder Type',
        yaxis_title='Count',
        template='plotly_white',
        barmode='group',
        height=500
    )

    # ------------------ RISK vs IMPACT MATRIX ------------------
    risk_impact_data = {}
    for stakeholder in qs:
        risk = stakeholder.risk_level or 'medium'
        impact = stakeholder.impact_level or 'medium'
        key = (risk, impact)
        risk_impact_data[key] = risk_impact_data.get(key, 0) + 1

    # Prepare data for heatmap
    risk_levels = ['very_low', 'low', 'medium', 'high', 'very_high']
    impact_levels = ['very_low', 'low', 'medium', 'high', 'very_high']

    heatmap_data = []
    for risk in risk_levels:
        row = []
        for impact in impact_levels:
            row.append(risk_impact_data.get((risk, impact), 0))
        heatmap_data.append(row)

    fig_risk_impact = go.Figure(data=go.Heatmap(
        z=heatmap_data,
        x=[level.replace('_', ' ').title() for level in impact_levels],
        y=[level.replace('_', ' ').title() for level in risk_levels],
        colorscale='RdYlGn_r',  # Red for high risk/impact, Green for low
        hoverongaps=False,
        hovertemplate='Risk: %{y}<br>Impact: %{x}<br>Count: %{z}<extra></extra>',
        text=heatmap_data,
        texttemplate="%{z}",
        textfont={"size": 12}
    ))
    fig_risk_impact.update_layout(
        title='Risk vs Impact Matrix<br><sub>Number of stakeholders in each risk-impact category</sub>',
        xaxis_title='Impact Level',
        yaxis_title='Risk Level',
        template='plotly_white',
        height=500
    )

    # ------------------ ENGAGEMENT STRATEGY EFFECTIVENESS ------------------
    engagement_effectiveness = {}
    for stakeholder in qs:
        satisfaction_value = {
            'very_low': 1, 'low': 2, 'medium': 3, 'high': 4, 'very_high': 5
        }.get(stakeholder.satisfaction_level, 3)

        strategies = stakeholder.engagement_strategy or ['none']
        for strategy in strategies:
            label = strategy.replace('_', ' ').title()
            if label not in engagement_effectiveness:
                engagement_effectiveness[label] = {
                    'count': 0,
                    'total_satisfaction': 0,
                    'avg_engagement_score': 0,
                    'engagement_scores': []
                }
            engagement_effectiveness[label]['count'] += 1
            engagement_effectiveness[label]['total_satisfaction'] += satisfaction_value
            engagement_effectiveness[label]['engagement_scores'].append(float(stakeholder.engagement_priority_score))

    # Calculate averages
    for strategy in engagement_effectiveness.values():
        strategy['avg_satisfaction'] = strategy['total_satisfaction'] / strategy['count']
        strategy['avg_engagement_score'] = sum(strategy['engagement_scores']) / len(strategy['engagement_scores'])

    engagement_labels = list(engagement_effectiveness.keys())
    engagement_counts = [data['count'] for data in engagement_effectiveness.values()]
    engagement_satisfaction = [data['avg_satisfaction'] for data in engagement_effectiveness.values()]
    engagement_scores = [data['avg_engagement_score'] for data in engagement_effectiveness.values()]

    fig_engagement = go.Figure()
    fig_engagement.add_trace(go.Bar(
        name='Stakeholder Count',
        x=engagement_labels,
        y=engagement_counts,
        marker_color=color_palette[0],
        text=engagement_counts,
        textposition='auto',
    ))
    fig_engagement.add_trace(go.Scatter(
        name='Avg Satisfaction',
        x=engagement_labels,
        y=engagement_satisfaction,
        mode='lines+markers',
        line=dict(color=color_palette[1], width=3),
        yaxis='y2'
    ))
    fig_engagement.update_layout(
        title='Engagement Strategy Effectiveness<br><sub>Bars show usage count, line shows average satisfaction</sub>',
        xaxis_title='Engagement Strategy',
        yaxis=dict(title='Count', side='left'),
        yaxis2=dict(title='Average Satisfaction (1-5)', side='right', overlaying='y', range=[1, 5]),
        template='plotly_white',
        height=500
    )

    # ------------------ PRIORITY ANALYSIS ------------------
    priority_analysis = qs.values('priority').annotate(
        count=Count('id'),
        avg_engagement=Avg('engagement_priority_score'),
        requires_attention_count=Count('id', filter=Q(requires_attention=True)),
        key_stakeholder_count=Count('id', filter=Q(is_key_stakeholder=True))
    ).order_by('priority')

    priority_labels = [p['priority'].replace('_', ' ').title() for p in priority_analysis]
    priority_counts = [p['count'] for p in priority_analysis]
    priority_attention = [p['requires_attention_count'] for p in priority_analysis]
    priority_key = [p['key_stakeholder_count'] for p in priority_analysis]

    fig_priority = go.Figure()
    fig_priority.add_trace(go.Bar(
        name='Total Stakeholders',
        x=priority_labels,
        y=priority_counts,
        marker_color=color_palette[0],
        text=priority_counts,
        textposition='auto'
    ))
    fig_priority.add_trace(go.Bar(
        name='Require Attention',
        x=priority_labels,
        y=priority_attention,
        marker_color=color_palette[1],
        text=priority_attention,
        textposition='auto'
    ))
    fig_priority.add_trace(go.Bar(
        name='Key Stakeholders',
        x=priority_labels,
        y=priority_key,
        marker_color=color_palette[2],
        text=priority_key,
        textposition='auto'
    ))
    fig_priority.update_layout(
        title='Stakeholder Priority Analysis<br><sub>Breakdown by priority level with attention and key stakeholder flags</sub>',
        xaxis_title='Priority Level',
        yaxis_title='Count',
        template='plotly_white',
        barmode='group',
        height=500
    )

    # ------------------ RELATIONSHIP STATUS OVERVIEW ------------------
    # Calculate overdue engagements manually since it's a property
    from datetime import date, timedelta
    ninety_days_ago = date.today() - timedelta(days=90)

    relationship_data = qs.values('relationship_status').annotate(
        count=Count('id'),
        avg_satisfaction=Avg(Case(
            When(satisfaction_level='very_low', then=1),
            When(satisfaction_level='low', then=2),
            When(satisfaction_level='medium', then=3),
            When(satisfaction_level='high', then=4),
            When(satisfaction_level='very_high', then=5),
            default=3,
            output_field=IntegerField()
        ))
    ).order_by('relationship_status')

    # Calculate overdue engagements for each relationship status
    relationship_overdue = {}
    for status in relationship_data:
        status_qs = qs.filter(relationship_status=status['relationship_status'])
        overdue_count = 0
        for stakeholder in status_qs:
            if (stakeholder.last_engagement_date and
                    stakeholder.last_engagement_date < ninety_days_ago):
                overdue_count += 1
        relationship_overdue[status['relationship_status']] = overdue_count

    relationship_labels = [r['relationship_status'].title() for r in relationship_data]
    relationship_counts = [r['count'] for r in relationship_data]
    relationship_satisfaction = [r['avg_satisfaction'] or 0 for r in relationship_data]
    relationship_overdue_counts = [relationship_overdue.get(status, 0) for status in
                                   [r['relationship_status'] for r in relationship_data]]

    fig_relationship = go.Figure()
    fig_relationship.add_trace(go.Bar(
        name='Total Stakeholders',
        x=relationship_labels,
        y=relationship_counts,
        marker_color=color_palette[0],
        text=relationship_counts,
        textposition='auto'
    ))
    fig_relationship.add_trace(go.Bar(
        name='Overdue Engagement',
        x=relationship_labels,
        y=relationship_overdue_counts,
        marker_color=color_palette[1],
        text=relationship_overdue_counts,
        textposition='auto'
    ))
    fig_relationship.update_layout(
        title='Relationship Status Overview<br><sub>Total stakeholders and overdue engagements by relationship status</sub>',
        xaxis_title='Relationship Status',
        yaxis_title='Count',
        template='plotly_white',
        barmode='group',
        height=500
    )

    # ------------------ ENHANCED SUMMARY DATA ------------------
    total_stakeholders = qs.count()

    # Calculate percentages and averages with None handling
    internal_count = qs.filter(stakeholder_type='internal').count()
    external_count = qs.filter(stakeholder_type='external').count()
    interface_count = qs.filter(stakeholder_type='interface').count()

    # Risk analysis
    high_risk_count = qs.filter(risk_level__in=['high', 'very_high']).count()
    medium_risk_count = qs.filter(risk_level='medium').count()
    low_risk_count = qs.filter(risk_level__in=['low', 'very_low']).count()

    # Priority analysis
    high_priority_count = qs.filter(priority__in=['high', 'very_high']).count()
    key_stakeholder_count = qs.filter(is_key_stakeholder=True).count()
    attention_required_count = qs.filter(requires_attention=True).count()

    # Engagement analysis - calculate overdue manually
    overdue_engagement_count = 0
    for stakeholder in qs:
        if (stakeholder.last_engagement_date and
                stakeholder.last_engagement_date < ninety_days_ago):
            overdue_engagement_count += 1

    avg_engagement_score = qs.aggregate(avg=Avg('engagement_priority_score'))['avg'] or 0

    # Satisfaction analysis
    high_satisfaction_count = qs.filter(satisfaction_level__in=['high', 'very_high']).count()
    low_satisfaction_count = qs.filter(satisfaction_level__in=['low', 'very_low']).count()

    summary_data = {
        'total_stakeholders': total_stakeholders,
        'internal_count': internal_count,
        'external_count': external_count,
        'interface_count': interface_count,
        'internal_percentage': round((internal_count / total_stakeholders * 100) if total_stakeholders > 0 else 0, 1),
        'external_percentage': round((external_count / total_stakeholders * 100) if total_stakeholders > 0 else 0, 1),
        'high_risk_count': high_risk_count,
        'high_risk_percentage': round((high_risk_count / total_stakeholders * 100) if total_stakeholders > 0 else 0, 1),
        'medium_risk_percentage': round((medium_risk_count / total_stakeholders * 100) if total_stakeholders > 0 else 0,
                                        1),
        'low_risk_percentage': round((low_risk_count / total_stakeholders * 100) if total_stakeholders > 0 else 0, 1),
        'high_priority_count': high_priority_count,
        'key_stakeholder_count': key_stakeholder_count,
        'attention_required_count': attention_required_count,
        'attention_percentage': round(
            (attention_required_count / total_stakeholders * 100) if total_stakeholders > 0 else 0, 1),
        'overdue_engagement_count': overdue_engagement_count,
        'overdue_percentage': round(
            (overdue_engagement_count / total_stakeholders * 100) if total_stakeholders > 0 else 0, 1),
        'avg_engagement_score': round(avg_engagement_score, 1),
        'high_satisfaction_count': high_satisfaction_count,
        'high_satisfaction_percentage': round(
            (high_satisfaction_count / total_stakeholders * 100) if total_stakeholders > 0 else 0, 1),
        'low_satisfaction_count': low_satisfaction_count,
        'low_satisfaction_percentage': round(
            (low_satisfaction_count / total_stakeholders * 100) if total_stakeholders > 0 else 0, 1),
    }

    return render(request, 'stakeholder_list/graph.html', {
        'plot_html_type_count': fig_type_count.to_html(full_html=False),  # New stakeholder type count box
        'plot_html_category': fig_category.to_html(full_html=False),
        'plot_html_role': fig_role.to_html(full_html=False),
        'plot_html_type': fig_type.to_html(full_html=False),
        'plot_html_risk_impact': fig_risk_impact.to_html(full_html=False),
        'plot_html_engagement': fig_engagement.to_html(full_html=False),
        'plot_html_priority': fig_priority.to_html(full_html=False),
        'plot_html_relationship': fig_relationship.to_html(full_html=False),
        'summary_data': summary_data,
    })