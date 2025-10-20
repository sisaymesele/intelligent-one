from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from datetime import date
import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, PatternFill

from django.utils import timezone
from django.db.models.functions import TruncMonth

from django.db.models import Q, Count, Sum, Avg, Max, Min
from django.db.models.functions import TruncMonth, TruncYear
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator

# Plotly for charts
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from plotly.offline import plot
# Date/Time
from datetime import timedelta, datetime, date

from management_project.models import InitiativeReport, InitiativePlanning
from management_project.forms import InitiativeReportForm

# -------------------- LIST  --------------------

@login_required
def initiative_report_list(request):
    query = request.GET.get('search', '').strip()
    selected_focus_area = request.GET.get('initiative_focus_area', '').strip()
    page_number = request.GET.get('page', 1)

    # Base queryset: all reports for initiatives in the user's organization
    timelines = InitiativeReport.objects.filter(
        organization_name=request.user.organization_name
    ).select_related('initiative_planning')

    # Filter by Focus Area (through ForeignKey)
    if selected_focus_area:
        timelines = timelines.filter(
            initiative_planning__initiative_focus_area=selected_focus_area
        )

    # Search filter across InitiativePlanning fields
    if query:
        timelines = timelines.filter(
            Q(initiative_planning__initiative_name__icontains=query) |
            Q(initiative_planning__initiative_dimension__icontains=query) |
            Q(initiative_planning__initiative_focus_area__icontains=query)

        )

    # Order by report_date descending
    timelines = timelines.order_by('-report_date')

    # Pagination
    paginator = Paginator(timelines, 10)
    page_obj = paginator.get_page(page_number)

    # Distinct focus areas for dropdown filter
    focus_areas = InitiativePlanning.objects.filter(
        organization_name=request.user.organization_name
    ).values_list('initiative_focus_area', flat=True).distinct().order_by('initiative_focus_area')

    return render(request, 'initiative_report/list.html', {
        'timelines': page_obj,
        'page_obj': page_obj,
        'search_query': query,
        'focus_areas': focus_areas,
        'selected_focus_area': selected_focus_area,
    })
# -------------------- CREATE  --------------------


@login_required
def create_initiative_report(request):
    selected_initiative = request.GET.get('initiative_planning')  # Preselect if passed
    if request.method == 'POST':
        form = InitiativeReportForm(request.POST, request=request)
        if form.is_valid() and 'save' in request.POST:
            timeline = form.save(commit=False)
            timeline.organization_name = request.user.organization_name
            timeline.save()
            messages.success(request, "InitiativePlanning timeline created successfully!")
            return redirect('initiative_report_list')  # or keep in child list page
    else:
        initial_data = {}
        if selected_initiative:
            initial_data['initiative_planning'] = selected_initiative
        form = InitiativeReportForm(initial=initial_data, request=request)

    return render(request, 'initiative_report/form.html', {
        'form': form,
        'next': None  # child form doesn't need next
    })


# -------------------- UPDATE --------------------
@login_required
def update_initiative_report(request, pk):
    timeline = get_object_or_404(
        InitiativeReport,
        pk=pk,
        organization_name=request.user.organization_name
    )

    if request.method == 'POST':
        form = InitiativeReportForm(request.POST, instance=timeline, request=request)
        if 'save' in request.POST and form.is_valid():
            form.save()
            messages.success(request, "InitiativePlanning timeline updated successfully!")
            return redirect('initiative_report_list')
        else:
            messages.error(request, "Error updating timeline. Please check the form.")
    else:
        form = InitiativeReportForm(instance=timeline, request=request)

    return render(request, 'initiative_report/form.html', {'form': form})


# -------------------- DELETE  --------------------
@login_required
def delete_initiative_report(request, pk):
    timeline = get_object_or_404(
        InitiativeReport,
        pk=pk,
        organization_name=request.user.organization_name
    )

    if request.method == 'POST':
        timeline.delete()
        messages.success(request, "InitiativePlanning timeline deleted successfully!")
        return redirect('initiative_report_list')

    return render(request, 'initiative_report/delete_confirm.html', {'timeline': timeline})




@login_required
def export_initiative_report_to_excel(request):
    """
    Export InitiativeReport queryset to Excel with colored title and header row.
    Supports filtering by focus area and search query.
    Includes all fields: budget, HR, status, dates, risk, and notes.
    """
    query = request.GET.get('search', '').strip()
    selected_focus_area = request.GET.get('initiative_focus_area', '').strip()

    # Base queryset for user's organization
    reports = InitiativeReport.objects.filter(
        organization_name=request.user.organization_name
    ).select_related('initiative_planning')

    # Apply filters
    if selected_focus_area:
        reports = reports.filter(
            initiative_planning__initiative_focus_area=selected_focus_area
        )

    if query:
        reports = reports.filter(
            Q(initiative_planning__initiative_name__icontains=query) |
            Q(initiative_planning__initiative_dimension__icontains=query)
        )

    reports = reports.order_by('-report_date')

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Initiative Reports"

    # ===== Title Row =====
    total_columns = 26  # update to match all columns
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws['A1']
    title_cell.value = "Initiative Performance Reports"
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # ===== Header Row =====
    headers = [
        "Focus Area", "Dimension", "Initiative Name", "Organization", "Report Date",
        "Planned Budget", "Budget Spent", "Budget Remaining", "Budget Utilization (%)",
        "Actual HR", "Remaining HR",
        "Baseline Status", "Target Status", "Achieved Status", "Status Achievement (%)",
        "Priority", "Impact", "Likelihood", "Risk Level",
        "Start Date", "End Date", "Remaining Days", "Remaining Months",
        "Notes"
    ]
    ws.append(headers)

    # Style header row
    for cell in ws[2]:  # row 2 after title
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # ===== Fill Data Rows =====
    for r in reports:
        row = [
            r.initiative_planning.initiative_focus_area,
            r.initiative_planning.initiative_dimension,
            r.initiative_planning.initiative_name,
            str(r.organization_name),
            r.report_date.strftime("%Y-%m-%d"),
            float(r.planned_budget),
            float(r.total_budget_spent),
            float(r.budget_remaining),
            float(r.budget_utilization_percent),
            float(r.total_actual_hr),
            float(r.remaining_hr),
            r.initiative_planning.baseline_status,
            r.initiative_planning.target_status,
            r.achieved_status,
            float(r.status_achievement_percent),
            r.initiative_planning.priority,
            r.initiative_planning.impact,
            r.initiative_planning.likelihood,
            r.initiative_planning.risk_level,
            r.initiative_planning.start_date.strftime("%Y-%m-%d") if r.initiative_planning.start_date else "",
            r.initiative_planning.end_date.strftime("%Y-%m-%d") if r.initiative_planning.end_date else "",
            r.remaining_days or "",
            r.remaining_months or "",
            r.notes or "",
        ]
        ws.append(row)

    # ===== Adjust Column Widths =====
    for column_cells in ws.iter_cols(min_row=2):
        non_merged_cells = [cell for cell in column_cells if not isinstance(cell, openpyxl.cell.cell.MergedCell)]
        if not non_merged_cells:
            continue
        length = max(len(str(cell.value)) if cell.value else 0 for cell in non_merged_cells)
        ws.column_dimensions[non_merged_cells[0].column_letter].width = min(length + 2, 50)

    # ===== Create Response =====
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=initiative_reports.xlsx'
    wb.save(response)
    return response

# views.py


@login_required
def initiative_report_charts(request):
    """Complete initiative dashboard with all charts and KPIs in single view"""

    # Get filter parameters
    query = request.GET.get('search', '').strip()
    selected_focus_area = request.GET.get('initiative_focus_area', '').strip()
    time_range = request.GET.get('time_range', 'all')

    # Base queryset with optimal database queries
    reports = InitiativeReport.objects.filter(
        organization_name=request.user.organization_name
    ).select_related('initiative_planning')

    # Apply time range filter
    if time_range != 'all':
        today = timezone.now().date()
        if time_range == '3m':
            start_date = today - timedelta(days=90)
        elif time_range == '6m':
            start_date = today - timedelta(days=180)
        elif time_range == '1y':
            start_date = today - timedelta(days=365)
        reports = reports.filter(report_date__gte=start_date)

    # Apply search filters
    if selected_focus_area:
        reports = reports.filter(initiative_planning__initiative_focus_area=selected_focus_area)

    if query:
        reports = reports.filter(
            Q(initiative_planning__initiative_name__icontains=query) |
            Q(initiative_planning__initiative_dimension__icontains=query) |
            Q(initiative_planning__initiative_focus_area__icontains=query)
        )

    # Get focus areas for dropdown
    focus_areas = InitiativePlanning.objects.filter(
        organization_name=request.user.organization_name
    ).values_list('initiative_focus_area', flat=True).distinct().order_by('initiative_focus_area')

    # Calculate comprehensive statistics
    total_reports = reports.count()
    reports_list = list(reports)

    # Calculate KPIs using model properties
    total_budget_planned = sum(float(report.planned_budget) for report in reports_list)
    total_budget_spent = sum(float(report.total_budget_spent) for report in reports_list)
    total_hr_planned = sum(float(report.planned_hr) for report in reports_list)
    total_hr_used = sum(float(report.total_actual_hr) for report in reports_list)

    budget_utilizations = [report.budget_utilization_percent for report in reports_list if report.planned_budget > 0]
    hr_utilizations = [report.hr_utilization_percent for report in reports_list if report.planned_hr > 0]
    achievement_values = [report.status_achievement_percent for report in reports_list]

    avg_budget_utilization = sum(budget_utilizations) / len(budget_utilizations) if budget_utilizations else 0
    avg_hr_utilization = sum(hr_utilizations) / len(hr_utilizations) if hr_utilizations else 0
    avg_achievement = sum(achievement_values) / len(achievement_values) if achievement_values else 0

    remaining_budget = total_budget_planned - total_budget_spent
    remaining_hr = total_hr_planned - total_hr_used

    # Performance distribution
    high_performers = sum(1 for r in reports_list if r.status_achievement_percent >= 80)
    moderate_performers = sum(1 for r in reports_list if 60 <= r.status_achievement_percent < 80)
    low_performers = sum(1 for r in reports_list if r.status_achievement_percent < 60)

    # Status distribution
    status_distribution = reports.values('achieved_status').annotate(
        count=Count('id')
    ).order_by('achieved_status')

    # Monthly trends for report count
    monthly_trends = reports.annotate(
        month=TruncMonth('report_date')
    ).values('month').annotate(
        report_count=Count('id')
    ).order_by('month')

    # Prepare stats for template
    stats = {
        'total_reports': total_reports,
        'total_budget_planned': total_budget_planned,
        'total_budget_spent': total_budget_spent,
        'remaining_budget': remaining_budget,
        'total_hr_planned': total_hr_planned,
        'total_hr_used': total_hr_used,
        'remaining_hr': remaining_hr,
        'avg_budget_utilization': avg_budget_utilization,
        'avg_hr_utilization': avg_hr_utilization,
        'avg_achievement': avg_achievement,
    }

    performance_distribution = {
        'high_performers': high_performers,
        'moderate_performers': moderate_performers,
        'low_performers': low_performers,
        'high_percent': round((high_performers / total_reports * 100), 1) if total_reports > 0 else 0,
        'moderate_percent': round((moderate_performers / total_reports * 100), 1) if total_reports > 0 else 0,
        'low_percent': round((low_performers / total_reports * 100), 1) if total_reports > 0 else 0,
    }

    # Generate all charts within the main function
    charts = {}

    # Common axis configuration
    axis_config = dict(
        showline=True,
        linewidth=2,
        linecolor='black',
        mirror=True,
        showgrid=True,
        gridwidth=1,
        gridcolor='lightgray',
        zeroline=True,
        zerolinewidth=2,
        zerolinecolor='gray'
    )

    # Chart 1: Budget Planned vs Actual
    if reports_list:
        initiative_names = [report.initiative_planning.initiative_name for report in reports_list]
        planned_budgets = [float(report.planned_budget) for report in reports_list]
        actual_spent = [float(report.total_budget_spent) for report in reports_list]

        fig1 = go.Figure()
        fig1.add_trace(go.Bar(
            name='Planned Budget', x=initiative_names, y=planned_budgets,
            marker_color='#4361ee', opacity=0.8,
            hovertemplate='<b>%{x}</b><br>Planned: $%{y:,.0f}<extra></extra>'
        ))
        fig1.add_trace(go.Bar(
            name='Actual Spent', x=initiative_names, y=actual_spent,
            marker_color='#f72585', opacity=0.8,
            hovertemplate='<b>%{x}</b><br>Actual: $%{y:,.0f}<extra></extra>'
        ))
        fig1.update_layout(
            title=dict(text='Budget: Planned vs Actual', x=0.5, xanchor='center', font=dict(size=16)),
            xaxis=dict(
                title='Initiatives',
                tickangle=45,
                **axis_config,
                title_font=dict(size=14, color='black'),
                tickfont=dict(size=12, color='black')
            ),
            yaxis=dict(
                title='Amount ($)',
                tickprefix='$',
                **axis_config,
                title_font=dict(size=14, color='black'),
                tickfont=dict(size=12, color='black')
            ),
            barmode='group',
            height=400,
            showlegend=True,
            legend=dict(
                orientation='h',
                yanchor='bottom',
                y=1.02,
                xanchor='right',
                x=1,
                font=dict(size=12)
            ),
            paper_bgcolor='white',
            plot_bgcolor='white',
            margin=dict(l=60, r=40, t=80, b=120)
        )
        charts['budget_chart'] = plot(fig1, output_type='div', include_plotlyjs=False)
    else:
        charts['budget_chart'] = create_empty_chart("No budget data available", axis_config)

    # Chart 2: HR Planned vs Actual
    if reports_list:
        initiative_names = [report.initiative_planning.initiative_name for report in reports_list]
        planned_hr = [float(report.planned_hr) for report in reports_list]
        actual_hr = [float(report.total_actual_hr) for report in reports_list]

        fig2 = go.Figure()
        fig2.add_trace(go.Bar(
            name='Planned HR', x=initiative_names, y=planned_hr,
            marker_color='#4cc9f0', opacity=0.8,
            hovertemplate='<b>%{x}</b><br>Planned: %{y:.1f} person-days<extra></extra>'
        ))
        fig2.add_trace(go.Bar(
            name='Actual HR', x=initiative_names, y=actual_hr,
            marker_color='#7209b7', opacity=0.8,
            hovertemplate='<b>%{x}</b><br>Actual: %{y:.1f} person-days<extra></extra>'
        ))
        fig2.update_layout(
            title=dict(text='HR: Planned vs Actual (Person-Days)', x=0.5, xanchor='center', font=dict(size=16)),
            xaxis=dict(
                title='Initiatives',
                tickangle=45,
                **axis_config,
                title_font=dict(size=14, color='black'),
                tickfont=dict(size=12, color='black')
            ),
            yaxis=dict(
                title='Person-Days',
                **axis_config,
                title_font=dict(size=14, color='black'),
                tickfont=dict(size=12, color='black')
            ),
            barmode='group',
            height=400,
            showlegend=True,
            legend=dict(
                orientation='h',
                yanchor='bottom',
                y=1.02,
                xanchor='right',
                x=1,
                font=dict(size=12)
            ),
            paper_bgcolor='white',
            plot_bgcolor='white',
            margin=dict(l=60, r=40, t=80, b=120)
        )
        charts['hr_chart'] = plot(fig2, output_type='div', include_plotlyjs=False)
    else:
        charts['hr_chart'] = create_empty_chart("No HR data available", axis_config)

    # Chart 3: Status Achievement %
    if reports_list:
        initiative_names = [report.initiative_planning.initiative_name for report in reports_list]
        achievement_percent = [report.status_achievement_percent for report in reports_list]

        # Sort by achievement percentage for better visualization
        sorted_data = sorted(zip(initiative_names, achievement_percent), key=lambda x: x[1])
        sorted_names, sorted_achievement = zip(*sorted_data)

        fig3 = go.Figure()
        fig3.add_trace(go.Scatter(
            x=list(sorted_names), y=list(sorted_achievement),
            mode='lines+markers',
            line=dict(color='#16a34a', width=3),
            marker=dict(size=8, color='#16a34a'),
            hovertemplate='<b>%{x}</b><br>Achievement: %{y:.1f}%<extra></extra>'
        ))
        fig3.add_hline(
            y=80,
            line_dash="dash",
            line_color="red",
            annotation_text="Target (80%)",
            annotation_position="bottom right",
            annotation_font=dict(size=12, color='red')
        )
        fig3.update_layout(
            title=dict(text='Status Achievement %', x=0.5, xanchor='center', font=dict(size=16)),
            xaxis=dict(
                title='Initiatives',
                tickangle=45,
                **axis_config,
                title_font=dict(size=14, color='black'),
                tickfont=dict(size=12, color='black')
            ),
            yaxis=dict(
                title='Achievement (%)',
                range=[0, 100],
                **axis_config,
                title_font=dict(size=14, color='black'),
                tickfont=dict(size=12, color='black')
            ),
            height=400,
            paper_bgcolor='white',
            plot_bgcolor='white',
            margin=dict(l=60, r=40, t=80, b=120)
        )
        charts['achievement_chart'] = plot(fig3, output_type='div', include_plotlyjs=False)
    else:
        charts['achievement_chart'] = create_empty_chart("No achievement data available", axis_config)

    # Chart 4: Budget Utilization %
    if reports_list:
        initiative_names = [report.initiative_planning.initiative_name for report in reports_list]
        utilization_percent = [report.budget_utilization_percent for report in reports_list]

        # Sort by utilization for better visualization
        sorted_data = sorted(zip(initiative_names, utilization_percent), key=lambda x: x[1])
        sorted_names, sorted_utilization = zip(*sorted_data)

        # Color based on utilization level
        colors = []
        for util in sorted_utilization:
            if util <= 80:
                colors.append('#16a34a')  # Green
            elif util <= 90:
                colors.append('#d97706')  # Orange
            else:
                colors.append('#dc2626')  # Red

        fig4 = go.Figure()
        fig4.add_trace(go.Bar(
            y=list(sorted_names), x=list(sorted_utilization), orientation='h',
            marker_color=colors,
            hovertemplate='<b>%{y}</b><br>Utilization: %{x:.1f}%<extra></extra>'
        ))
        fig4.add_vline(
            x=80,
            line_dash="dash",
            line_color="blue",
            annotation_text="Optimal (80%)",
            annotation_position="top right",
            annotation_font=dict(size=12, color='blue')
        )
        fig4.update_layout(
            title=dict(text='Budget Utilization %', x=0.5, xanchor='center', font=dict(size=16)),
            yaxis=dict(
                title='Initiatives',
                **axis_config,
                title_font=dict(size=14, color='black'),
                tickfont=dict(size=12, color='black')
            ),
            xaxis=dict(
                title='Utilization (%)',
                range=[0, 100],
                **axis_config,
                title_font=dict(size=14, color='black'),
                tickfont=dict(size=12, color='black')
            ),
            height=400,
            paper_bgcolor='white',
            plot_bgcolor='white',
            margin=dict(l=80, r=40, t=80, b=80)
        )
        charts['budget_utilization_chart'] = plot(fig4, output_type='div', include_plotlyjs=False)
    else:
        charts['budget_utilization_chart'] = create_empty_chart("No budget utilization data available", axis_config)

    # Chart 5: HR Utilization %
    if reports_list:
        initiative_names = [report.initiative_planning.initiative_name for report in reports_list]
        utilization_percent = [report.hr_utilization_percent for report in reports_list]

        # Sort by utilization for better visualization
        sorted_data = sorted(zip(initiative_names, utilization_percent), key=lambda x: x[1])
        sorted_names, sorted_utilization = zip(*sorted_data)

        # Color based on utilization level
        colors = []
        for util in sorted_utilization:
            if util <= 80:
                colors.append('#4cc9f0')  # Blue
            elif util <= 90:
                colors.append('#f72585')  # Pink
            else:
                colors.append('#7209b7')  # Purple

        fig5 = go.Figure()
        fig5.add_trace(go.Bar(
            y=list(sorted_names), x=list(sorted_utilization), orientation='h',
            marker_color=colors,
            hovertemplate='<b>%{y}</b><br>HR Utilization: %{x:.1f}%<extra></extra>'
        ))
        fig5.add_vline(
            x=80,
            line_dash="dash",
            line_color="blue",
            annotation_text="Optimal (80%)",
            annotation_position="top right",
            annotation_font=dict(size=12, color='blue')
        )
        fig5.update_layout(
            title=dict(text='HR Utilization %', x=0.5, xanchor='center', font=dict(size=16)),
            yaxis=dict(
                title='Initiatives',
                **axis_config,
                title_font=dict(size=14, color='black'),
                tickfont=dict(size=12, color='black')
            ),
            xaxis=dict(
                title='Utilization (%)',
                range=[0, 100],
                **axis_config,
                title_font=dict(size=14, color='black'),
                tickfont=dict(size=12, color='black')
            ),
            height=400,
            paper_bgcolor='white',
            plot_bgcolor='white',
            margin=dict(l=80, r=40, t=80, b=80)
        )
        charts['hr_utilization_chart'] = plot(fig5, output_type='div', include_plotlyjs=False)
    else:
        charts['hr_utilization_chart'] = create_empty_chart("No HR utilization data available", axis_config)

    # Chart 6: Monthly Trend of Reports
    if monthly_trends:
        months = [trend['month'].strftime('%b %Y') for trend in monthly_trends]
        report_counts = [trend['report_count'] for trend in monthly_trends]

        fig6 = go.Figure()
        fig6.add_trace(go.Scatter(
            x=months, y=report_counts, mode='lines+markers',
            line=dict(color='#7c3aed', width=3),
            marker=dict(size=8, color='#7c3aed'),
            fill='tozeroy',
            fillcolor='rgba(124, 58, 237, 0.1)',
            hovertemplate='<b>%{x}</b><br>Reports: %{y}<extra></extra>'
        ))
        fig6.update_layout(
            title=dict(text='Monthly Trend of Reports', x=0.5, xanchor='center', font=dict(size=16)),
            xaxis=dict(
                title='Month',
                tickangle=45,
                **axis_config,
                title_font=dict(size=14, color='black'),
                tickfont=dict(size=12, color='black')
            ),
            yaxis=dict(
                title='Number of Reports',
                **axis_config,
                title_font=dict(size=14, color='black'),
                tickfont=dict(size=12, color='black')
            ),
            height=400,
            paper_bgcolor='white',
            plot_bgcolor='white',
            margin=dict(l=60, r=40, t=80, b=120)
        )
        charts['monthly_trend_chart'] = plot(fig6, output_type='div', include_plotlyjs=False)
    else:
        charts['monthly_trend_chart'] = create_empty_chart("No monthly trend data available", axis_config)

    context = {
        'charts': charts,
        'focus_areas': focus_areas,
        'selected_focus_area': selected_focus_area,
        'search_query': query,
        'time_range': time_range,
        'stats': stats,
        'status_distribution': status_distribution,
        'performance_distribution': performance_distribution,
        'total_reports': total_reports,
        'initiative_count': len(set(r.initiative_planning_id for r in reports_list)),
        'monthly_trends': list(monthly_trends),
    }

    return render(request, 'initiative_report/chart.html', context)


def create_empty_chart(message, axis_config):
    """Create an empty chart with message and visible axes"""
    fig = go.Figure()
    fig.add_annotation(
        text=message,
        xref="paper",
        yref="paper",
        x=0.5,
        y=0.5,
        showarrow=False,
        font=dict(size=16, color="gray", family="Arial")
    )
    fig.update_layout(
        height=400,
        paper_bgcolor='white',
        plot_bgcolor='white',
        xaxis=dict(
            **axis_config,
            title='X Axis',
            title_font=dict(size=14, color='black'),
            tickfont=dict(size=12, color='black')
        ),
        yaxis=dict(
            **axis_config,
            title='Y Axis',
            title_font=dict(size=14, color='black'),
            tickfont=dict(size=12, color='black')
        ),
        margin=dict(l=60, r=40, t=80, b=80)
    )
    return plot(fig, output_type='div', include_plotlyjs=False)
#
# @login_required
# def initiative_report_charts(request):
#     """Complete initiative dashboard with all charts and KPIs in single view"""
#
#     # Get filter parameters
#     query = request.GET.get('search', '').strip()
#     selected_focus_area = request.GET.get('initiative_focus_area', '').strip()
#     time_range = request.GET.get('time_range', 'all')
#
#     # Base queryset with optimal database queries
#     reports = InitiativeReport.objects.filter(
#         organization_name=request.user.organization_name
#     ).select_related('initiative_planning')
#
#     # Apply time range filter
#     if time_range != 'all':
#         today = timezone.now().date()
#         if time_range == '3m':
#             start_date = today - timedelta(days=90)
#         elif time_range == '6m':
#             start_date = today - timedelta(days=180)
#         elif time_range == '1y':
#             start_date = today - timedelta(days=365)
#         reports = reports.filter(report_date__gte=start_date)
#
#     # Apply search filters
#     if selected_focus_area:
#         reports = reports.filter(initiative_planning__initiative_focus_area=selected_focus_area)
#
#     if query:
#         reports = reports.filter(
#             Q(initiative_planning__initiative_name__icontains=query) |
#             Q(initiative_planning__initiative_dimension__icontains=query) |
#             Q(initiative_planning__initiative_focus_area__icontains=query)
#         )
#
#     # Get focus areas for dropdown
#     focus_areas = InitiativePlanning.objects.filter(
#         organization_name=request.user.organization_name
#     ).values_list('initiative_focus_area', flat=True).distinct().order_by('initiative_focus_area')
#
#     # Calculate comprehensive statistics
#     total_reports = reports.count()
#     reports_list = list(reports)
#
#     # Calculate KPIs using model properties
#     total_budget_planned = sum(float(report.planned_budget) for report in reports_list)
#     total_budget_spent = sum(float(report.total_budget_spent) for report in reports_list)
#     total_hr_planned = sum(float(report.planned_hr) for report in reports_list)
#     total_hr_used = sum(float(report.total_actual_hr) for report in reports_list)
#
#     budget_utilizations = [report.budget_utilization_percent for report in reports_list if report.planned_budget > 0]
#     hr_utilizations = [report.hr_utilization_percent for report in reports_list if report.planned_hr > 0]
#     achievement_values = [report.status_achievement_percent for report in reports_list]
#
#     avg_budget_utilization = sum(budget_utilizations) / len(budget_utilizations) if budget_utilizations else 0
#     avg_hr_utilization = sum(hr_utilizations) / len(hr_utilizations) if hr_utilizations else 0
#     avg_achievement = sum(achievement_values) / len(achievement_values) if achievement_values else 0
#
#     remaining_budget = total_budget_planned - total_budget_spent
#     remaining_hr = total_hr_planned - total_hr_used
#
#     # Performance distribution
#     high_performers = sum(1 for r in reports_list if r.status_achievement_percent >= 80)
#     moderate_performers = sum(1 for r in reports_list if 60 <= r.status_achievement_percent < 80)
#     low_performers = sum(1 for r in reports_list if r.status_achievement_percent < 60)
#
#     # Status distribution
#     status_distribution = reports.values('achieved_status').annotate(
#         count=Count('id')
#     ).order_by('achieved_status')
#
#     # Monthly trends for report count
#     monthly_trends = reports.annotate(
#         month=TruncMonth('report_date')
#     ).values('month').annotate(
#         report_count=Count('id')
#     ).order_by('month')
#
#     # Prepare stats for template
#     stats = {
#         'total_reports': total_reports,
#         'total_budget_planned': total_budget_planned,
#         'total_budget_spent': total_budget_spent,
#         'remaining_budget': remaining_budget,
#         'total_hr_planned': total_hr_planned,
#         'total_hr_used': total_hr_used,
#         'remaining_hr': remaining_hr,
#         'avg_budget_utilization': avg_budget_utilization,
#         'avg_hr_utilization': avg_hr_utilization,
#         'avg_achievement': avg_achievement,
#     }
#
#     performance_distribution = {
#         'high_performers': high_performers,
#         'moderate_performers': moderate_performers,
#         'low_performers': low_performers,
#         'high_percent': round((high_performers / total_reports * 100), 1) if total_reports > 0 else 0,
#         'moderate_percent': round((moderate_performers / total_reports * 100), 1) if total_reports > 0 else 0,
#         'low_percent': round((low_performers / total_reports * 100), 1) if total_reports > 0 else 0,
#     }
#
#     # Generate all charts within the main function
#     charts = {}
#
#     # Chart 1: Budget Planned vs Actual
#     if reports_list:
#         initiative_names = [report.initiative_planning.initiative_name for report in reports_list]
#         planned_budgets = [float(report.planned_budget) for report in reports_list]
#         actual_spent = [float(report.total_budget_spent) for report in reports_list]
#
#         fig1 = go.Figure()
#         fig1.add_trace(go.Bar(
#             name='Planned Budget', x=initiative_names, y=planned_budgets,
#             marker_color='#4361ee', opacity=0.8,
#             hovertemplate='<b>%{x}</b><br>Planned: $%{y:,.0f}<extra></extra>'
#         ))
#         fig1.add_trace(go.Bar(
#             name='Actual Spent', x=initiative_names, y=actual_spent,
#             marker_color='#f72585', opacity=0.8,
#             hovertemplate='<b>%{x}</b><br>Actual: $%{y:,.0f}<extra></extra>'
#         ))
#         fig1.update_layout(
#             title=dict(text='Budget: Planned vs Actual', x=0.5, xanchor='center'),
#             xaxis=dict(title='Initiatives', tickangle=45),
#             yaxis=dict(title='Amount ($)', tickprefix='$'),
#             barmode='group', height=400, showlegend=True,
#             legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
#             paper_bgcolor='white', plot_bgcolor='white'
#         )
#         charts['budget_chart'] = plot(fig1, output_type='div', include_plotlyjs=False)
#     else:
#         charts['budget_chart'] = create_empty_chart("No budget data available")
#
#     # Chart 2: HR Planned vs Actual
#     if reports_list:
#         initiative_names = [report.initiative_planning.initiative_name for report in reports_list]
#         planned_hr = [float(report.planned_hr) for report in reports_list]
#         actual_hr = [float(report.total_actual_hr) for report in reports_list]
#
#         fig2 = go.Figure()
#         fig2.add_trace(go.Bar(
#             name='Planned HR', x=initiative_names, y=planned_hr,
#             marker_color='#4cc9f0', opacity=0.8,
#             hovertemplate='<b>%{x}</b><br>Planned: %{y:.1f} person-days<extra></extra>'
#         ))
#         fig2.add_trace(go.Bar(
#             name='Actual HR', x=initiative_names, y=actual_hr,
#             marker_color='#7209b7', opacity=0.8,
#             hovertemplate='<b>%{x}</b><br>Actual: %{y:.1f} person-days<extra></extra>'
#         ))
#         fig2.update_layout(
#             title=dict(text='HR: Planned vs Actual (Person-Days)', x=0.5, xanchor='center'),
#             xaxis=dict(title='Initiatives', tickangle=45),
#             yaxis=dict(title='Person-Days'), barmode='group', height=400, showlegend=True,
#             legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
#             paper_bgcolor='white', plot_bgcolor='white'
#         )
#         charts['hr_chart'] = plot(fig2, output_type='div', include_plotlyjs=False)
#     else:
#         charts['hr_chart'] = create_empty_chart("No HR data available")
#
#     # Chart 3: Status Achievement %
#     if reports_list:
#         initiative_names = [report.initiative_planning.initiative_name for report in reports_list]
#         achievement_percent = [report.status_achievement_percent for report in reports_list]
#
#         # Sort by achievement percentage for better visualization
#         sorted_data = sorted(zip(initiative_names, achievement_percent), key=lambda x: x[1])
#         sorted_names, sorted_achievement = zip(*sorted_data)
#
#         fig3 = go.Figure()
#         fig3.add_trace(go.Scatter(
#             x=list(sorted_names), y=list(sorted_achievement),
#             mode='lines+markers', line=dict(color='#16a34a', width=3),
#             marker=dict(size=8, color='#16a34a'),
#             hovertemplate='<b>%{x}</b><br>Achievement: %{y:.1f}%<extra></extra>'
#         ))
#         fig3.add_hline(y=80, line_dash="dash", line_color="red",
#                        annotation_text="Target (80%)", annotation_position="bottom right")
#         fig3.update_layout(
#             title=dict(text='Status Achievement %', x=0.5, xanchor='center'),
#             xaxis=dict(title='Initiatives', tickangle=45),
#             yaxis=dict(title='Achievement (%)', range=[0, 100]), height=400,
#             paper_bgcolor='white', plot_bgcolor='white'
#         )
#         charts['achievement_chart'] = plot(fig3, output_type='div', include_plotlyjs=False)
#     else:
#         charts['achievement_chart'] = create_empty_chart("No achievement data available")
#
#     # Chart 4: Budget Utilization %
#     if reports_list:
#         initiative_names = [report.initiative_planning.initiative_name for report in reports_list]
#         utilization_percent = [report.budget_utilization_percent for report in reports_list]
#
#         # Sort by utilization for better visualization
#         sorted_data = sorted(zip(initiative_names, utilization_percent), key=lambda x: x[1])
#         sorted_names, sorted_utilization = zip(*sorted_data)
#
#         # Color based on utilization level
#         colors = []
#         for util in sorted_utilization:
#             if util <= 80:
#                 colors.append('#16a34a')  # Green
#             elif util <= 90:
#                 colors.append('#d97706')  # Orange
#             else:
#                 colors.append('#dc2626')  # Red
#
#         fig4 = go.Figure()
#         fig4.add_trace(go.Bar(
#             y=list(sorted_names), x=list(sorted_utilization), orientation='h',
#             marker_color=colors,
#             hovertemplate='<b>%{y}</b><br>Utilization: %{x:.1f}%<extra></extra>'
#         ))
#         fig4.add_vline(x=80, line_dash="dash", line_color="blue",
#                        annotation_text="Optimal (80%)", annotation_position="top right")
#         fig4.update_layout(
#             title=dict(text='Budget Utilization %', x=0.5, xanchor='center'),
#             yaxis=dict(title='Initiatives'), xaxis=dict(title='Utilization (%)', range=[0, 100]),
#             height=400, paper_bgcolor='white', plot_bgcolor='white'
#         )
#         charts['budget_utilization_chart'] = plot(fig4, output_type='div', include_plotlyjs=False)
#     else:
#         charts['budget_utilization_chart'] = create_empty_chart("No budget utilization data available")
#
#     # Chart 5: HR Utilization %
#     if reports_list:
#         initiative_names = [report.initiative_planning.initiative_name for report in reports_list]
#         utilization_percent = [report.hr_utilization_percent for report in reports_list]
#
#         # Sort by utilization for better visualization
#         sorted_data = sorted(zip(initiative_names, utilization_percent), key=lambda x: x[1])
#         sorted_names, sorted_utilization = zip(*sorted_data)
#
#         # Color based on utilization level
#         colors = []
#         for util in sorted_utilization:
#             if util <= 80:
#                 colors.append('#4cc9f0')  # Blue
#             elif util <= 90:
#                 colors.append('#f72585')  # Pink
#             else:
#                 colors.append('#7209b7')  # Purple
#
#         fig5 = go.Figure()
#         fig5.add_trace(go.Bar(
#             y=list(sorted_names), x=list(sorted_utilization), orientation='h',
#             marker_color=colors,
#             hovertemplate='<b>%{y}</b><br>HR Utilization: %{x:.1f}%<extra></extra>'
#         ))
#         fig5.add_vline(x=80, line_dash="dash", line_color="blue",
#                        annotation_text="Optimal (80%)", annotation_position="top right")
#         fig5.update_layout(
#             title=dict(text='HR Utilization %', x=0.5, xanchor='center'),
#             yaxis=dict(title='Initiatives'), xaxis=dict(title='Utilization (%)', range=[0, 100]),
#             height=400, paper_bgcolor='white', plot_bgcolor='white'
#         )
#         charts['hr_utilization_chart'] = plot(fig5, output_type='div', include_plotlyjs=False)
#     else:
#         charts['hr_utilization_chart'] = create_empty_chart("No HR utilization data available")
#
#     # Chart 6: Monthly Trend of Reports
#     if monthly_trends:
#         months = [trend['month'].strftime('%b %Y') for trend in monthly_trends]
#         report_counts = [trend['report_count'] for trend in monthly_trends]
#
#         fig6 = go.Figure()
#         fig6.add_trace(go.Scatter(
#             x=months, y=report_counts, mode='lines+markers',
#             line=dict(color='#7c3aed', width=3), marker=dict(size=8, color='#7c3aed'),
#             fill='tozeroy', fillcolor='rgba(124, 58, 237, 0.1)',
#             hovertemplate='<b>%{x}</b><br>Reports: %{y}<extra></extra>'
#         ))
#         fig6.update_layout(
#             title=dict(text='Monthly Trend of Reports', x=0.5, xanchor='center'),
#             xaxis=dict(title='Month', tickangle=45), yaxis=dict(title='Number of Reports'),
#             height=400, paper_bgcolor='white', plot_bgcolor='white'
#         )
#         charts['monthly_trend_chart'] = plot(fig6, output_type='div', include_plotlyjs=False)
#     else:
#         charts['monthly_trend_chart'] = create_empty_chart("No monthly trend data available")
#
#     context = {
#         'charts': charts,
#         'focus_areas': focus_areas,
#         'selected_focus_area': selected_focus_area,
#         'search_query': query,
#         'time_range': time_range,
#         'stats': stats,
#         'status_distribution': status_distribution,
#         'performance_distribution': performance_distribution,
#         'total_reports': total_reports,
#         'initiative_count': len(set(r.initiative_planning_id for r in reports_list)),
#         'monthly_trends': list(monthly_trends),
#     }
#
#     return render(request, 'initiative_report/chart.html', context)
#
#
# def create_empty_chart(message):
#     """Create an empty chart with message and visible axes"""
#     fig = go.Figure()
#     fig.add_annotation(
#         text=message, xref="paper", yref="paper",
#         x=0.5, y=0.5, showarrow=False,
#         font=dict(size=16, color="gray")
#     )
#     fig.update_layout(
#         height=400,
#         paper_bgcolor='white',
#         plot_bgcolor='white',
#         xaxis=dict(
#             showgrid=True,
#             zeroline=True,
#             showticklabels=True,
#             gridcolor='lightgray',
#             zerolinecolor='gray'
#         ),
#         yaxis=dict(
#             showgrid=True,
#             zeroline=True,
#             showticklabels=True,
#             gridcolor='lightgray',
#             zerolinecolor='gray'
#         )
#     )
#     return plot(fig, output_type='div', include_plotlyjs=False)

