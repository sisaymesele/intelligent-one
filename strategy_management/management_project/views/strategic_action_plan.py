from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.urls import reverse
from django.contrib import messages
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from datetime import date
import calendar
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from django.shortcuts import render
import plotly.graph_objs as go
from plotly.offline import plot
from collections import defaultdict

from management_project.models import StrategicCycle, StrategicActionPlan, Stakeholder
from management_project.forms import StrategicActionPlanForm


@login_required
def strategic_action_plan_by_cycle(request):
    """List distinct strategic cycles for the current organization with all info."""
    cycles_qs = StrategicCycle.objects.filter(
        organization_name=request.user.organization_name
    ).order_by('-start_date')

    # Build a list of dicts including calculated properties
    cycles = []
    for cycle in cycles_qs:
        cycles.append({
            'name': cycle.name,
            'time_horizon': cycle.time_horizon,
            'time_horizon_type': cycle.time_horizon_type,
            'start_date': cycle.start_date,
            'end_date': cycle.end_date,
            'slug': cycle.slug,
            'duration_days': (cycle.end_date - cycle.start_date).days if cycle.start_date and cycle.end_date else None,
            'start_month_name': calendar.month_name[cycle.start_date.month] if cycle.start_date else None,
            'start_quarter': (cycle.start_date.month - 1) // 3 + 1 if cycle.start_date else None,
            'start_year': cycle.start_date.year if cycle.start_date else None,
        })

    return render(request, 'strategic_action_plan/cycle_list.html', {
        'strategic_cycles': cycles
    })


@login_required
def strategic_action_plan_list(request, cycle_slug):
    strategy_by_cycle = get_object_or_404(
        StrategicCycle,
        slug=cycle_slug,
        organization_name=request.user.organization_name
    )

    # Base queryset
    strategic_action_plans = StrategicActionPlan.objects.filter(
        strategic_cycle=strategy_by_cycle,
        organization_name=request.user.organization_name
    ).order_by('-id')

    # Search query
    search_query = request.GET.get('search', '').strip()

    if search_query:
        strategic_action_plans = strategic_action_plans.filter(
            Q(strategy_hierarchy__objective__icontains=search_query) |
            Q(strategy_hierarchy__kpi__icontains=search_query) |
            Q(strategy_hierarchy__strategic_perspective__icontains=search_query) |
            Q(strategy_hierarchy__focus_area__icontains=search_query) |
            Q(strategic_cycle__time_horizon__icontains=search_query) |
            Q(strategic_cycle__time_horizon_type__icontains=search_query) |
            Q(responsible_bodies__stakeholder_name__icontains=search_query)
        ).distinct()

    # Pagination
    paginator = Paginator(strategic_action_plans, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Form for creating new action plan
    form = StrategicActionPlanForm(initial={'strategic_cycle': strategy_by_cycle}, request=request)

    context = {
        'strategy_by_cycle': strategy_by_cycle,
        'page_obj': page_obj,
        'form': form,
        'search_query': search_query,
    }

    return render(request, 'strategic_action_plan/list_by_cycle.html', context)




@login_required
def strategic_action_plan_detail(request, cycle_slug, pk):
    strategy_by_cycle = get_object_or_404(StrategicCycle, slug=cycle_slug)
    strategic_action_plan = get_object_or_404(
        StrategicActionPlan,
        pk=pk,
        strategic_cycle=strategy_by_cycle,
        organization_name=request.user.organization_name
    )

    return render(request, 'strategic_action_plan/detail.html', {
        'strategy_by_cycle': strategy_by_cycle,
        'strategic_action_plan': strategic_action_plan
    })
#
#
# @login_required
# def create_strategic_action_plan(request, cycle_slug):
#     strategy_by_cycle = get_object_or_404(StrategicCycle, slug=cycle_slug)
#     next_url = request.GET.get('next') or request.POST.get('next')
#
#     # Preselect parents if passed
#     selected_cycle = request.GET.get('cycle') or request.POST.get('cycle')
#     selected_strategy = request.GET.get('strategy_hierarchy') or request.POST.get('strategy_hierarchy')
#     selected_responsible = request.GET.getlist('responsible_bodies') or request.POST.getlist('responsible_bodies')
#
#     if request.method == "POST":
#         form = StrategicActionPlanForm(request.POST, request=request)
#         if form.is_valid() and 'save' in request.POST:
#             sap = form.save(commit=False)
#
#             # Assign strategic cycle if selected
#             if selected_cycle:
#                 sap.strategic_cycle_id = selected_cycle
#             else:
#                 sap.strategic_cycle = strategy_by_cycle
#
#             sap.organization_name = request.user.organization_name
#             sap.save()
#             form.save_m2m()
#
#             messages.success(request, "Strategic action plan saved successfully!")
#
#             # Redirect back if 'next' provided
#             if next_url:
#                 separator = '&' if '?' in next_url else '?'
#                 return redirect(f"{next_url}{separator}sap={sap.pk}")
#
#             return redirect("strategic_action_plan_list", cycle_slug=strategy_by_cycle.slug)
#
#     else:
#         # Set initial preselected values
#         initial_data = {}
#         if selected_cycle:
#             initial_data['strategic_cycle'] = selected_cycle
#         if selected_strategy:
#             initial_data['strategy_hierarchy'] = selected_strategy
#         if selected_responsible:
#             initial_data['responsible_bodies'] = selected_responsible
#
#         form = StrategicActionPlanForm(initial=initial_data, request=request)
#
#     return render(request, "strategic_action_plan/form.html", {
#         "form": form,
#         "form_title": f"{'Update' if form.instance.pk else 'Create'} Strategic Action Plan",
#         "submit_button_text": "Save",
#         "back_url": reverse("strategic_action_plan_list", kwargs={"cycle_slug": strategy_by_cycle.slug}),
#         "strategy_by_cycle": strategy_by_cycle,
#         "next": next_url,
#     })

@login_required
def create_strategic_action_plan(request, cycle_slug):
    strategy_by_cycle = get_object_or_404(StrategicCycle, slug=cycle_slug)
    next_url = request.GET.get('next') or request.POST.get('next')

    selected_strategy = request.GET.get('strategy_hierarchy') or request.POST.get('strategy_hierarchy')
    selected_responsible = request.GET.getlist('responsible_bodies') or request.POST.getlist('responsible_bodies')

    if request.method == "POST":
        form = StrategicActionPlanForm(request.POST, request=request, cycle=strategy_by_cycle)
        if form.is_valid() and 'save' in request.POST:
            sap = form.save(commit=False)
            sap.strategic_cycle = strategy_by_cycle  # auto-assign from URL
            sap.organization_name = request.user.organization_name
            sap.save()
            form.save_m2m()
            messages.success(request, "Strategic action plan saved successfully!")

            if next_url:
                separator = '&' if '?' in next_url else '?'
                return redirect(f"{next_url}{separator}sap={sap.pk}")

            return redirect("strategic_action_plan_list", cycle_slug=strategy_by_cycle.slug)

    else:
        initial_data = {}
        if selected_strategy:
            initial_data['strategy_hierarchy'] = selected_strategy
        if selected_responsible:
            initial_data['responsible_bodies'] = selected_responsible

        form = StrategicActionPlanForm(initial=initial_data, request=request, cycle=strategy_by_cycle)

    return render(request, "strategic_action_plan/form.html", {
        "form": form,
        "form_title": f"{'Update' if form.instance.pk else 'Create'} Strategic Action Plan",
        "submit_button_text": "Save",
        "back_url": reverse("strategic_action_plan_list", kwargs={"cycle_slug": strategy_by_cycle.slug}),
        "strategy_by_cycle": strategy_by_cycle,
        "next": next_url,
    })




@login_required
def update_strategic_action_plan(request, cycle_slug, pk):
    strategy_by_cycle = get_object_or_404(StrategicCycle, slug=cycle_slug)
    sap = get_object_or_404(
        StrategicActionPlan,
        pk=pk,
        strategic_cycle=strategy_by_cycle,
        organization_name=request.user.organization_name
    )

    if request.method == 'POST':
        form = StrategicActionPlanForm(request.POST, instance=sap, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, "Strategic action plan updated successfully!")
            return redirect('strategic_action_plan_list', cycle_slug=strategy_by_cycle.slug)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = StrategicActionPlanForm(instance=sap, request=request)

    return render(request, 'strategic_action_plan/form.html', {
        'form': form,
        'form_title': f'Update Strategic Action Plan for {strategy_by_cycle.name}',
        'submit_button_text': 'Update Strategic Action Plan',
        'back_url': reverse('strategic_action_plan_list', kwargs={'cycle_slug': strategy_by_cycle.slug}),
        'strategy_by_cycle': strategy_by_cycle,
        'edit_strategic_action_plan': sap
    })


@login_required
def delete_strategic_action_plan(request, cycle_slug, pk):
    strategy_by_cycle = get_object_or_404(StrategicCycle, slug=cycle_slug)
    sap = get_object_or_404(
        StrategicActionPlan,
        pk=pk,
        strategic_cycle=strategy_by_cycle,
        organization_name=request.user.organization_name
    )

    if request.method == 'POST':
        sap.delete()
        messages.success(request, "Strategic action plan deleted successfully!")
        return redirect('strategic_action_plan_list', cycle_slug=strategy_by_cycle.slug)

    return render(request, 'strategic_action_plan/delete_confirm.html', {
        'strategic_action_plan': sap,
        'strategy_by_cycle': strategy_by_cycle
    })




def split_two_lines(text):
    """Split a long text roughly into two lines."""
    if not text:
        return ""
    words = text.split()
    mid = len(words) // 2
    return " ".join(words[:mid]) + "\n" + " ".join(words[mid:])

@login_required
def export_strategic_action_plan_to_excel(request, cycle_slug):
    cycle = get_object_or_404(StrategicCycle, slug=cycle_slug, organization_name=request.user.organization_name)
    plans = StrategicActionPlan.objects.filter(strategic_cycle=cycle, organization_name=request.user.organization_name)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"{cycle.name} ({cycle.start_date.year if cycle.start_date else ''})"

    # Title row
    title_text = f"Strategic Action Plan For: {cycle.name}"
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=17)
    title_cell = sheet.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    title_cell.border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    # Column headers
    headers = [
        "#", "Perspective", "Pillar", "Objective", "KPI", "Indicator Type",
        "Direction", "Baseline", "Target", "Improvement Needed",
        "Time Horizon", "Time Horizon Type", "Start Date", "End Date",
        "Duration (Days)", "Responsible Bodies", "Status"
    ]
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid")
        cell.border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

    # Data rows
    for row_idx, plan in enumerate(plans, start=3):
        data = [
            row_idx - 2,
            plan.strategy_hierarchy.strategic_perspective if plan.strategy_hierarchy else "",
            split_two_lines(plan.strategy_hierarchy.focus_area if plan.strategy_hierarchy else ""),
            split_two_lines(plan.strategy_hierarchy.objective if plan.strategy_hierarchy else ""),
            split_two_lines(plan.strategy_hierarchy.kpi if plan.strategy_hierarchy else ""),
            plan.get_indicator_type_display(),
            plan.get_direction_of_change_display(),
            plan.baseline,
            plan.target,
            plan.improvement_needed or 0.0,
            plan.strategic_cycle.time_horizon if plan.strategic_cycle else "",
            plan.strategic_cycle.time_horizon_type if plan.strategic_cycle else "",
            plan.strategic_cycle.start_date.strftime('%B %d, %Y') if plan.strategic_cycle and plan.strategic_cycle.start_date else "",
            plan.strategic_cycle.end_date.strftime('%B %d, %Y') if plan.strategic_cycle and plan.strategic_cycle.end_date else "",
            plan.strategic_cycle.duration_days if plan.strategic_cycle else "",
            ", ".join([body.stakeholder_name for body in plan.responsible_bodies.all()]),
            plan.get_status_display() or "Pending",
        ]
        for col_num, value in enumerate(data, start=1):
            cell = sheet.cell(row=row_idx, column=col_num, value=value)
            # Wrap text for certain fields
            if col_num in [3, 4, 5, 16]:  # Pillar, Objective, KPI, Responsible Bodies
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            # Borders
            cell.border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )

    # Minimized column widths with a reasonable minimum
    min_width = 10
    max_width = 20
    for col_idx in range(1, sheet.max_column + 1):
        width = min_width
        for row_idx in range(2, sheet.max_row + 1):  # skip title row
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                width = max(width, min(len(str(cell.value)) + 2, max_width))
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    # Save to buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Strategic_Action_Plan_{cycle.name}.xlsx"'
    return response


#
def strategic_action_plan_chart(request):

    # ---------------- Base queryset ----------------
    base_qs = (
        StrategicActionPlan.objects.select_related("strategy_hierarchy", "strategic_cycle")
        .prefetch_related("responsible_bodies")
        .filter(organization_name=request.user.organization_name)  # 🔹 restrict to user organization
    )

    # ---------------- Filter values ----------------
    selected_cycle = request.GET.get("cycle")
    selected_body = request.GET.get("body")

    plans = base_qs
    if selected_cycle:
        plans = plans.filter(strategic_cycle__id=selected_cycle)
    if selected_body:
        plans = plans.filter(responsible_bodies__id=selected_body)

    # ---------------- Unique sorted lists ----------------
    cycles = StrategicCycle.objects.filter(organization_name=request.user.organization_name).order_by("name")
    responsible_bodies = Stakeholder.objects.filter(organization_name=request.user.organization_name).order_by("stakeholder_name")

    # ---------------- Aggregate function ----------------
    def aggregate(plans_subset):
        objective_counts, status_data, perspective_data, pillar_data = (
            defaultdict(int), defaultdict(int), defaultdict(int), defaultdict(int)
        )
        for plan in plans_subset:
            objective = plan.strategy_hierarchy.objective or "Uncategorized"
            status = plan.get_status_display()
            perspective = plan.strategy_hierarchy.strategic_perspective or "Uncategorized"
            pillar = plan.strategy_hierarchy.focus_area or "Uncategorized"

            objective_counts[objective] += 1
            status_data[status] += 1
            perspective_data[perspective] += 1
            pillar_data[pillar] += 1
        return objective_counts, status_data, perspective_data, pillar_data

    # ---------------- Color Schemes ----------------
    STATUS_COLORS = {
        "Pending": "#FFB347",      # orange
        "In Progress": "#1f77b4",  # blue
        "Completed": "#2ca02c",    # green
        "On Hold": "#9467bd",      # purple
        "Cancelled": "#d62728",    # red
    }
    CATEGORY_COLORS = ["#17BECF", "#BCBD22", "#1f77b4", "#FF7F0E", "#8C564B"]

    # ---------------- Overall Charts ----------------
    obj_all, status_all, pers_all, pillar_all = aggregate(base_qs)

    fig_overall_pie = go.Figure(go.Pie(
        labels=list(status_all.keys()),
        values=list(status_all.values()),
        hole=0.4,
        textinfo="label+percent",
        marker=dict(colors=[STATUS_COLORS.get(s, "#7f7f7f") for s in status_all.keys()])
    ))
    fig_overall_pie.update_layout(title="Status Distribution (Overall)", height=400)

    categories_all = sorted(set(list(pers_all) + list(pillar_all) + list(obj_all)))
    fig_overall_bar = go.Figure()
    fig_overall_bar.add_trace(go.Bar(x=categories_all, y=[pers_all.get(c,0) for c in categories_all],
                                     name="Perspective", marker_color=CATEGORY_COLORS[0]))
    fig_overall_bar.add_trace(go.Bar(x=categories_all, y=[pillar_all.get(c,0) for c in categories_all],
                                     name="Pillar", marker_color=CATEGORY_COLORS[1]))
    fig_overall_bar.add_trace(go.Bar(x=categories_all, y=[obj_all.get(c,0) for c in categories_all],
                                     name="Objective", marker_color=CATEGORY_COLORS[2]))
    fig_overall_bar.update_layout(
        title="Plans by Perspective, Pillar, and Objective (Overall)",
        barmode="group", height=500, xaxis_tickangle=-45
    )

    # ---------------- Filtered Charts ----------------
    obj_filtered, status_filtered, pers_filtered, pillar_filtered = aggregate(plans)

    fig_filtered_pie = go.Figure(go.Pie(
        labels=list(status_filtered.keys()),
        values=list(status_filtered.values()),
        hole=0.4,
        textinfo="label+percent",
        marker=dict(colors=[STATUS_COLORS.get(s, "#7f7f7f") for s in status_filtered.keys()])
    ))
    fig_filtered_pie.update_layout(title="Status Distribution (Filtered)", height=400)

    categories_filtered = sorted(set(list(pers_filtered) + list(pillar_filtered) + list(obj_filtered)))
    fig_filtered_bar = go.Figure()
    fig_filtered_bar.add_trace(go.Bar(x=categories_filtered, y=[pers_filtered.get(c,0) for c in categories_filtered],
                                      name="Perspective", marker_color=CATEGORY_COLORS[0]))
    fig_filtered_bar.add_trace(go.Bar(x=categories_filtered, y=[pillar_filtered.get(c,0) for c in categories_filtered],
                                      name="Pillar", marker_color=CATEGORY_COLORS[1]))
    fig_filtered_bar.add_trace(go.Bar(x=categories_filtered, y=[obj_filtered.get(c,0) for c in categories_filtered],
                                      name="Objective", marker_color=CATEGORY_COLORS[2]))
    fig_filtered_bar.update_layout(
        title="Plans by Perspective, Pillar, and Objective (Filtered)",
        barmode="group", height=500, xaxis_tickangle=-45
    )

    # ---------------- KPI Summary ----------------
    def kpi_summary(obj, status):
        total = sum(obj.values())
        completed = status.get("Completed", 0)
        rate = (completed / total * 100) if total else 0
        return total, completed, round(rate, 2)

    total_all, completed_all, rate_all = kpi_summary(obj_all, status_all)
    total_filtered, completed_filtered, rate_filtered = kpi_summary(obj_filtered, status_filtered)

    # ---------------- Context ----------------
    context = {
        # Charts
        "plot_overall_pie": plot(fig_overall_pie, output_type="div", include_plotlyjs=True),
        "plot_overall_bar": plot(fig_overall_bar, output_type="div", include_plotlyjs=False),
        "plot_filtered_pie": plot(fig_filtered_pie, output_type="div", include_plotlyjs=False),
        "plot_filtered_bar": plot(fig_filtered_bar, output_type="div", include_plotlyjs=False),

        # KPIs
        "total_plans_all": total_all,
        "completed_all": completed_all,
        "completion_rate_all": rate_all,
        "total_plans_filtered": total_filtered,
        "completed_filtered": completed_filtered,
        "completion_rate_filtered": rate_filtered,

        # Dropdowns
        "cycles": cycles,
        "responsible_bodies": responsible_bodies,
        "selected_cycle": int(selected_cycle) if selected_cycle else None,
        "selected_body": int(selected_body) if selected_body else None,
    }
    return render(request, "strategic_action_plan/chart.html", context)
