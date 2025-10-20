from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from management_project.models import SwotAnalysis
from management_project.forms import SwotAnalysisForm
from django.db.models import Q

from django.core.paginator import Paginator

from django.http import HttpResponse
from django.core.paginator import Paginator

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Avg, Case, When, IntegerField
import plotly.graph_objects as go

from management_project.models import SwotAnalysis


# -------------------- SWOT LIST --------------------

@login_required
def swot_analysis_list(request):
    query = request.GET.get('search', '').strip()
    selected_type = request.GET.get('swot_type', '').strip()
    page_number = request.GET.get('page', 1)

    # Base queryset: all SWOT analyses for this organization
    swots = SwotAnalysis.objects.filter(
        organization_name=request.user.organization_name
    )

    # Filter by SWOT type
    if selected_type:
        swots = swots.filter(swot_type=selected_type)

    # Search filter across multiple fields
    if query:
        swots = swots.filter(
            Q(swot_type__icontains=query) |
            Q(swot_pillar__icontains=query) |
            Q(swot_factor__icontains=query) |
            Q(description__icontains=query)
        )

    # Ordering
    swots = swots.order_by('swot_type', 'priority', '-created_at')

    # Provide SWOT types for dropdown filter
    swot_types = [choice[0] for choice in SwotAnalysis.SWOT_TYPES]

    # Pagination
    paginator = Paginator(swots, 10)
    page_obj = paginator.get_page(page_number)

    return render(request, 'swot_analysis/list.html', {
        'swots': page_obj,
        'page_obj': page_obj,
        'search_query': query,
        'swot_types': swot_types,
        'selected_type': selected_type,
    })

# -------------------- CREATE SWOT --------------------


@login_required
def create_swot_analysis(request):
    if request.method == 'POST':
        form = SwotAnalysisForm(request.POST)
        # Only save if the Save button is clicked
        if 'save' in request.POST and form.is_valid():
            swot_entry = form.save(commit=False)
            swot_entry.organization_name = request.user.organization_name
            swot_entry.save()
            messages.success(request, "SWOT entry created successfully!")
            return redirect('swot_analysis_list')
    else:
        form = SwotAnalysisForm()

    return render(request, 'swot_analysis/form.html', {'form': form})


# -------------------- UPDATE SWOT --------------------
@login_required
def update_swot_analysis(request, pk):
    # Fetch the entry only if it belongs to the user's organization
    entry = get_object_or_404(SwotAnalysis, pk=pk, organization_name=request.user.organization_name)

    if request.method == 'POST':
        form = SwotAnalysisForm(request.POST, instance=entry)
        if 'save' in request.POST and form.is_valid():
            form.save()
            messages.success(request, "SWOT entry updated successfully!")
            return redirect('swot_analysis_list')
        else:
            messages.error(request, "Error updating SWOT entry. Please check the form.")
    else:
        form = SwotAnalysisForm(instance=entry)

    return render(request, 'swot_analysis/form.html', {'form': form})


# -------------------- DELETE SWOT --------------------

@login_required
def delete_swot_analysis(request, pk):
    entry = get_object_or_404(
        SwotAnalysis,
        pk=pk,
        organization_name=request.user.organization_name
    )

    if request.method == 'POST':
        entry.delete()
        messages.success(request, "SWOT entry deleted successfully!")
        return redirect('swot_analysis_list')

    return render(request, 'swot_analysis/delete_confirm.html', {'entry': entry})



@login_required
def export_swot_analysis_to_excel(request):
    # Get filters from request
    query = request.GET.get('search', '').strip()
    selected_type = request.GET.get('swot_type', '').strip()

    # Base queryset
    swots = SwotAnalysis.objects.filter(
        organization_name=request.user.organization_name
    )

    # Apply filters
    if selected_type:
        swots = swots.filter(swot_type=selected_type)
    if query:
        search_filter = (
            Q(swot_type__icontains=query) |
            Q(swot_pillar__icontains=query) |
            Q(swot_factor__icontains=query) |
            Q(description__icontains=query)
        )
        swots = swots.filter(search_filter)

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SWOT Analysis"

    # Styles
    title_font = Font(size=14, bold=True, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_fill = PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title row
    total_columns = 9
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "SWOT Analysis Report"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = title_fill

    # Header row
    headers = [
        "SWOT Type", "Pillar", "Factor", "Description",
        "Priority", "Impact", "Likelihood",
        "Created At", "Updated At"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Data rows
    for row_num, swot in enumerate(swots, start=3):
        row_data = [
            swot.swot_type,
            swot.swot_pillar,
            swot.swot_factor,
            swot.description or "",
            swot.priority,
            swot.impact,
            swot.likelihood or "",
            swot.created_at.strftime("%Y-%m-%d %H:%M"),
            swot.updated_at.strftime("%Y-%m-%d %H:%M"),
        ]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    # Auto-adjust column widths (skip merged cells)
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = max_length + 5

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=SWOT_Analysis_Report.xlsx'
    wb.save(response)
    return response



@login_required
def swot_analysis_chart(request):
    qs = SwotAnalysis.objects.filter(
        organization_name=request.user.organization_name
    )

    # ------------------ SWOT TYPE DISTRIBUTION ------------------
    type_counts = qs.values('swot_type').annotate(count=Count('id'))
    type_labels = [t['swot_type'] for t in type_counts]
    type_values = [t['count'] for t in type_counts]

    fig_swot_type = go.Figure(data=[go.Bar(
        x=type_labels,
        y=type_values,
        text=type_values,
        textposition='auto',
        marker_color=['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728']  # consistent colors
    )])
    fig_swot_type.update_layout(
        title='SWOT Distribution by Type',
        xaxis_title='SWOT Type',
        yaxis_title='Count',
        template='plotly_white',
        height=400
    )

    # ------------------ PRIORITY DISTRIBUTION ------------------
    priority_counts = qs.values('priority').annotate(count=Count('id'))
    priority_labels = [p['priority'] for p in priority_counts]
    priority_values = [p['count'] for p in priority_counts]

    fig_priority = go.Figure(data=[go.Bar(
        x=priority_labels,
        y=priority_values,
        text=priority_values,
        textposition='auto',
        marker_color=['#d62728', '#ff7f0e', '#ffbb78', '#98df8a', '#2ca02c']
    )])
    fig_priority.update_layout(
        title='SWOT Distribution by Priority',
        xaxis_title='Priority',
        yaxis_title='Count',
        template='plotly_white',
        height=400
    )

    # ------------------ IMPACT DISTRIBUTION ------------------
    impact_counts = qs.values('impact').annotate(count=Count('id'))
    impact_labels = [i['impact'] for i in impact_counts]
    impact_values = [i['count'] for i in impact_counts]

    fig_impact = go.Figure(data=[go.Bar(
        x=impact_labels,
        y=impact_values,
        text=impact_values,
        textposition='auto',
        marker_color=['#98df8a', '#2ca02c', '#ffbb78', '#ff7f0e', '#d62728']
    )])
    fig_impact.update_layout(
        title='SWOT Distribution by Impact',
        xaxis_title='Impact',
        yaxis_title='Count',
        template='plotly_white',
        height=400
    )

    # ------------------ LIKELIHOOD DISTRIBUTION ------------------
    likelihood_counts = qs.values('likelihood').annotate(count=Count('id'))
    likelihood_labels = [l['likelihood'] if l['likelihood'] else 'Unknown' for l in likelihood_counts]
    likelihood_values = [l['count'] for l in likelihood_counts]

    fig_likelihood = go.Figure(data=[go.Bar(
        x=likelihood_labels,
        y=likelihood_values,
        text=likelihood_values,
        textposition='auto',
        marker_color=['#17becf', '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#7f7f7f']
    )])
    fig_likelihood.update_layout(
        title='SWOT Distribution by Likelihood',
        xaxis_title='Likelihood',
        yaxis_title='Count',
        template='plotly_white',
        height=400
    )

    # ------------------ PILLAR ANALYSIS ------------------
    pillar_counts = qs.values('swot_pillar').annotate(count=Count('id')).order_by('-count')
    pillar_labels = [p['swot_pillar'] for p in pillar_counts]
    pillar_values = [p['count'] for p in pillar_counts]

    fig_pillar = go.Figure(data=[go.Bar(
        x=pillar_labels,
        y=pillar_values,
        text=pillar_values,
        textposition='auto',
        marker_color=['#636efa', '#ef553b', '#00cc96', '#ab63fa', '#ffa15a']
    )])
    fig_pillar.update_layout(
        title='SWOT Count per Pillar',
        xaxis_title='Pillar',
        yaxis_title='Count',
        template='plotly_white',
        height=400
    )

    # ------------------ SUMMARY DATA ------------------
    total_swot = qs.count()
    high_priority_count = qs.filter(priority__in=['High', 'Very High']).count()
    strengths_count = qs.filter(swot_type='Strength').count()
    weaknesses_count = qs.filter(swot_type='Weakness').count()
    opportunities_count = qs.filter(swot_type='Opportunity').count()
    threats_count = qs.filter(swot_type='Threat').count()

    summary_data = {
        'total_swot': total_swot,
        'high_priority_count': high_priority_count,
        'strengths_count': strengths_count,
        'weaknesses_count': weaknesses_count,
        'opportunities_count': opportunities_count,
        'threats_count': threats_count,
    }

    return render(request, 'swot_analysis/chart.html', {
        'plot_html_swot_type': fig_swot_type.to_html(full_html=False),
        'plot_html_priority': fig_priority.to_html(full_html=False),
        'plot_html_impact': fig_impact.to_html(full_html=False),
        'plot_html_likelihood': fig_likelihood.to_html(full_html=False),
        'plot_html_pillar': fig_pillar.to_html(full_html=False),
        'summary_data': summary_data,
    })

