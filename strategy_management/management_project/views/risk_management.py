from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q

from management_project.models import RiskManagement, StrategicCycle
from management_project.forms import RiskManagementForm

from io import BytesIO
from django.http import HttpResponse
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter



# -------------------- LIST (Category-based) --------------------
@login_required
def risk_management_list(request):
    search_query = request.GET.get('search', '').strip()
    selected_cycle = request.GET.get('strategic_cycle', '').strip()
    page_number = request.GET.get('page', 1)

    # Filter risks for the logged-in user's organization
    risks = RiskManagement.objects.filter(
        organization_name=request.user.organization_name
    )

    # Filter by strategic cycle
    if selected_cycle:
        risks = risks.filter(strategic_cycle__id=selected_cycle)

    # Apply search filter (name, category, or mitigation)
    if search_query:
        risks = risks.filter(
            Q(risk_category__icontains=search_query) |
            Q(risk_name__icontains=search_query) |
            Q(mitigation_action__icontains=search_query)
        )

    # Order by category (A-Z), then newest first
    risks = risks.order_by('risk_category', '-created_at')

    # Get Strategic Cycles for dropdown
    cycle_choices = StrategicCycle.objects.filter(
        organization_name=request.user.organization_name
    )

    # Paginate the results (10 per page)
    paginator = Paginator(risks, 10)
    page_obj = paginator.get_page(page_number)

    return render(request, 'risk_management/list.html', {
        'risks': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'cycle_choices': cycle_choices,
        'selected_cycle': selected_cycle,
    })


# -------------------- CREATE --------------------

@login_required
def create_risk_management(request):
    """
    Create a new RiskManagement entry.
    Allows cascading dropdowns to work without saving until Save is clicked.
    """
    if request.method == 'POST':
        # Pass request to form to allow filtering if needed
        form = RiskManagementForm(request.POST, request=request)

        # Only save if 'save' button was clicked
        if 'save' in request.POST:
            if form.is_valid():
                risk_entry = form.save(commit=False)
                risk_entry.organization_name = request.user.organization_name
                risk_entry.save()
                messages.success(request, "Risk entry created successfully!")
                return redirect('risk_management_list')
            else:
                messages.error(request, "Please fix the errors below.")
        # Else: just re-render the form (dropdown changed)
    else:
        form = RiskManagementForm(request=request)

    return render(request, 'risk_management/form.html', {'form': form})


# -------------------- UPDATE --------------------
# -------------------- UPDATE --------------------
@login_required
def update_risk_management(request, pk):
    """
    Update an existing RiskManagement entry.
    Cascading dropdowns still work without saving until Save is clicked.
    """
    entry = get_object_or_404(RiskManagement, pk=pk, organization_name=request.user.organization_name)

    if request.method == 'POST':
        form = RiskManagementForm(request.POST, instance=entry, request=request)

        # Only save if the Save button was clicked
        if 'save' in request.POST:
            if form.is_valid():
                form.save()
                messages.success(request, "Risk entry updated successfully!")
                return redirect('risk_management_list')
            else:
                messages.error(request, "Please fix the errors below.")
        # Else: just re-render the form for cascading dropdowns
    else:
        form = RiskManagementForm(instance=entry, request=request)

    return render(request, 'risk_management/form.html', {'form': form})




# -------------------- DELETE --------------------
@login_required
def delete_risk_management(request, pk):
    entry = get_object_or_404(RiskManagement, pk=pk, organization_name=request.user.organization_name)

    if request.method == 'POST':
        entry.delete()
        messages.success(request, "Risk entry deleted successfully!")
        return redirect('risk_management_list')

    return render(request, 'risk_management/delete_confirm.html', {'entry': entry})





def split_two_lines(text):
    """Split long text roughly into two lines for Excel display."""
    if not text:
        return ""
    words = text.split()
    mid = len(words) // 2
    return " ".join(words[:mid]) + "\n" + " ".join(words[mid:])

@login_required
def export_risk_management_excel(request):
    """Export risk management entries to Excel with strategic cycle filtering and search."""
    search_query = request.GET.get('search', '').strip()
    strategic_cycle_id = request.GET.get('strategic_cycle', '').strip()

    # Filter risks for the logged-in user's organization
    risks = RiskManagement.objects.filter(
        organization_name=request.user.organization_name
    )

    # Filter by selected strategic cycle
    if strategic_cycle_id:
        risks = risks.filter(strategic_cycle_id=strategic_cycle_id)

    # Apply search filter (category, name, mitigation)
    if search_query:
        risks = risks.filter(
            Q(risk_category__icontains=search_query) |
            Q(risk_name__icontains=search_query) |
            Q(mitigation_action__icontains=search_query)
        )

    # Order by category (A-Z), then newest created
    risks = risks.order_by('risk_category', '-created_at')

    # Create Excel workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Risk Management"

    # Title row
    title_text = "Risk Management Report"
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    title_cell = sheet.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    title_cell.border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    # Column headers
    headers = [
        "#", "Risk Category", "Risk Name", "Mitigation Action",
        "Likelihood", "Impact", "Severity Score", "Status",
        "Strategic Cycle", "Created At"
    ]
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid")
        cell.border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

    # Data rows
    for row_idx, risk in enumerate(risks, start=3):
        data = [
            row_idx - 2,
            split_two_lines(risk.risk_category),
            split_two_lines(risk.risk_name),
            split_two_lines(risk.mitigation_action),
            risk.get_likelihood_display(),
            risk.get_impact_display(),
            risk.severity_score,
            risk.get_status_display(),
            risk.strategic_cycle.name if risk.strategic_cycle else "",
            risk.created_at.strftime('%B %d, %Y %H:%M') if risk.created_at else "",
        ]

        for col_num, value in enumerate(data, start=1):
            cell = sheet.cell(row=row_idx, column=col_num, value=value)
            # Wrap text for long fields
            if col_num in [2, 3, 4, 9]:
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            # Borders
            cell.border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )

    # Adjust column widths
    min_width = 10
    max_width = 30
    for col_idx in range(1, sheet.max_column + 1):
        width = min_width
        for row_idx in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                width = max(width, min(len(str(cell.value)) + 2, max_width))
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    # Save workbook to buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Risk_Management_Report.xlsx"'
    return response

