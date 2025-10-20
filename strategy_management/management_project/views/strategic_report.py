from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.urls import reverse
from django.db.models import Q
import calendar

from django.contrib.auth.decorators import login_required
from management_project.models import StrategicReport, StrategicActionPlan, StrategicCycle
from management_project.forms import StrategicReportForm

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

#chart
from django.shortcuts import render
from django.db.models import Avg, Count
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from plotly.offline import plot
from collections import defaultdict
from datetime import datetime
from plotly.colors import qualitative


@login_required
def strategy_report_by_cycle_list(request):
    """List distinct strategic cycles for the current organization with all info."""
    cycles_qs = StrategicCycle.objects.filter(
        organization_name=request.user.organization_name
    ).order_by('-start_date')

    # Build a list of dicts including calculated properties
    cycles = []
    for cycle in cycles_qs:
        cycles.append({
            'name': cycle.name,
            'time_horizon': cycle.time_horizon,
            'time_horizon_type': cycle.time_horizon_type,
            'start_date': cycle.start_date,
            'end_date': cycle.end_date,
            'slug': cycle.slug,
            'duration_days': (cycle.end_date - cycle.start_date).days if cycle.start_date and cycle.end_date else None,
            'start_month_name': calendar.month_name[cycle.start_date.month] if cycle.start_date else None,
            'start_quarter': (cycle.start_date.month - 1) // 3 + 1 if cycle.start_date else None,
            'start_year': cycle.start_date.year if cycle.start_date else None,
        })

    return render(request, 'strategic_report/cycle_list.html', {
        'strategic_cycles': cycles
    })


@login_required
def strategic_report_list(request, cycle_slug):
    """List all strategic reports by cycle, grouped by action plan."""
    strategy_by_cycle = get_object_or_404(
        StrategicCycle,
        slug=cycle_slug,
        organization_name=request.user.organization_name
    )

    # Base queryset
    reports = StrategicReport.objects.filter(
        action_plan__strategic_cycle=strategy_by_cycle,
        organization_name=request.user.organization_name
    ).select_related("action_plan").order_by("-id")

    # Search query
    search_query = request.GET.get("search", "").strip()
    if search_query:
        reports = reports.filter(
            Q(action_plan__strategy_hierarchy__strategic_perspective__icontains=search_query) |
            Q(action_plan__strategy_hierarchy__focus_area__icontains=search_query) |
            Q(action_plan__strategy_hierarchy__objective__icontains=search_query) |
            Q(action_plan__strategy_hierarchy__kpi__icontains=search_query) |
            Q(action_plan__strategic_cycle__end_date__icontains=search_query) |
            Q(achievement__icontains=search_query) |
            Q(status__icontains=search_query)
        ).distinct()

    # Pagination
    paginator = Paginator(reports, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Form for creating new report
    form = StrategicReportForm(request=request, cycle=strategy_by_cycle)

    context = {
        "strategy_by_cycle": strategy_by_cycle,
        "page_obj": page_obj,
        "form": form,
        "search_query": search_query,
    }
    return render(request, "strategic_report/list_by_cycle.html", context)

@login_required
def strategic_report_detail(request, cycle_slug, pk):
    # Get the strategic cycle by slug
    strategy_by_cycle = get_object_or_404(StrategicCycle, slug=cycle_slug)

    # Get the specific strategic report for this cycle and organization
    strategic_report = get_object_or_404(
        StrategicReport,
        pk=pk,
        action_plan__strategic_cycle=strategy_by_cycle,
        organization_name=request.user.organization_name
    )

    return render(request, 'strategic_report/detail.html', {
        'strategy_by_cycle': strategy_by_cycle,
        'strategic_report': strategic_report
    })


@login_required
def create_strategic_report(request, cycle_slug):
    strategy_by_cycle = get_object_or_404(StrategicCycle, slug=cycle_slug)

    if request.method == "POST":
        form = StrategicReportForm(request.POST, request=request, cycle=strategy_by_cycle)
        if form.is_valid():
            report = form.save(commit=False)
            report.organization_name = request.user.organization_name
            report.save()
            form.save_m2m()
            messages.success(request, "Strategic report created successfully!")
            return redirect("strategic_report_list", cycle_slug=strategy_by_cycle.slug)
    else:
        form = StrategicReportForm(request=request, cycle=strategy_by_cycle)

    return render(request, "strategic_report/form.html", {
        "form": form,
        "form_title": f"Create Strategic Report for {strategy_by_cycle.name}",
        "submit_button_text": "Create Strategic Report",
        "back_url": reverse("strategic_report_list", kwargs={"cycle_slug": strategy_by_cycle.slug}),
        "strategy_by_cycle": strategy_by_cycle,
    })


@login_required
def update_strategic_report(request, cycle_slug, pk):
    strategy_by_cycle = get_object_or_404(StrategicCycle, slug=cycle_slug)
    report = get_object_or_404(
        StrategicReport,
        pk=pk,
        action_plan__strategic_cycle=strategy_by_cycle,
        organization_name=request.user.organization_name,
    )

    if request.method == "POST":
        form = StrategicReportForm(request.POST, instance=report, request=request, cycle=strategy_by_cycle)
        if form.is_valid():
            form.save()
            messages.success(request, "Strategic report updated successfully!")
            return redirect("strategic_report_list", cycle_slug=strategy_by_cycle.slug)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = StrategicReportForm(instance=report, request=request, cycle=strategy_by_cycle)

    return render(request, "strategic_report/form.html", {
        "form": form,
        "form_title": f"Update Strategic Report for {strategy_by_cycle.name}",
        "submit_button_text": "Update Strategic Report",
        "back_url": reverse("strategic_report_list", kwargs={"cycle_slug": strategy_by_cycle.slug}),
        "strategy_by_cycle": strategy_by_cycle,
        "edit_strategic_report": report,
    })


@login_required
def delete_strategic_report(request, cycle_slug, pk):
    strategy_by_cycle = get_object_or_404(StrategicCycle, slug=cycle_slug)
    report = get_object_or_404(
        StrategicReport,
        pk=pk,
        action_plan__strategic_cycle=strategy_by_cycle,
        organization_name=request.user.organization_name,
    )

    if request.method == "POST":
        report.delete()
        messages.success(request, "Strategic report deleted successfully!")
        return redirect("strategic_report_list", cycle_slug=strategy_by_cycle.slug)

    return render(request, "strategic_report/delete_confirm.html", {
        "strategic_report": report,
        "strategy_by_cycle": strategy_by_cycle,
    })



def break_text_every_3_words(text):
    """Break long text into new lines after every 3 words."""
    if not text:
        return ""
    words = text.split()
    return "\n".join(" ".join(words[i:i+3]) for i in range(0, len(words), 3))


@login_required
def export_strategic_report_to_excel(request, cycle_slug):
    """Export Strategic Reports with title, colored headers, and word breaks."""

    # 1️⃣ Get the cycle
    cycle = get_object_or_404(
        StrategicCycle,
        slug=cycle_slug,
        organization_name=request.user.organization_name,
    )

    # 2️⃣ Query reports
    reports = (
        StrategicReport.objects.filter(
            action_plan__strategic_cycle=cycle,
            organization_name=request.user.organization_name,
        )
        .select_related("organization_name", "action_plan", "action_plan__strategy_hierarchy")
        .prefetch_related("action_plan__responsible_bodies")
        .order_by("-id")
    )

    # 3️⃣ Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{cycle.name[:28]} Reports"

    # 4️⃣ Define styles
    total_columns = 26
    title_font = Font(size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid")
    border_style = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    # 5️⃣ Add title row (merged)
    title_text = f"Strategic Reports for {cycle.name}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # 6️⃣ Column headers
    headers = [
        "Report ID", "Organization", "Responsible Party", "Perspective",
        "Focus Area / Pillar", "Objective", "KPI", "Indicator Type",
        "Direction of Change", "Baseline", "Target", "Improvement Needed",
        "Achievement", "Percent Achieved", "Variance", "Weighted Score",
        "Data Source", "Data Collector", "Performance Summary",
        "Progress Summary", "Challenges", "Successes", "Lessons Learned",
        "Status", "Created At", "Updated At"
    ]
    ws.append(headers)

    # Format header row
    for cell in ws[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_style

    # 7️⃣ Add data rows
    for report in reports:
        ap = getattr(report, "action_plan", None)
        hierarchy = getattr(ap, "strategy_hierarchy", None)

        ws.append([
            report.id,
            getattr(report.organization_name, "organization_name", report.organization_name),
            ", ".join([body.stakeholder_name for body in ap.responsible_bodies.all()]) if ap else "-",
            getattr(hierarchy, "strategic_perspective", "-"),
            getattr(hierarchy, "focus_area", "-"),
            getattr(hierarchy, "objective", "-"),
            getattr(hierarchy, "kpi", "-"),
            getattr(ap, "indicator_type", "-") if ap else "-",
            getattr(ap, "direction_of_change", "-") if ap else "-",
            getattr(ap, "baseline", "-") if ap else "-",
            getattr(ap, "target", "-") if ap else "-",
            getattr(ap, "improvement_needed", "-") if ap else "-",
            getattr(report, "achievement", "-"),
            getattr(report, "percent_achieved", "-"),
            getattr(report, "variance", "-"),
            getattr(report, "weighted_score", "-"),
            break_text_every_3_words(getattr(report, "data_source", "")),
            break_text_every_3_words(getattr(report, "data_collector", "")),
            break_text_every_3_words(getattr(report, "performance_summary", "")),
            break_text_every_3_words(getattr(report, "progress_summary", "")),
            break_text_every_3_words(getattr(report, "challenges", "")),
            break_text_every_3_words(getattr(report, "successes", "")),
            break_text_every_3_words(getattr(report, "lessons_learned", "")),
            getattr(report, "get_status_display", lambda: "-")(),
            report.created_at.strftime("%Y-%m-%d %H:%M") if report.created_at else "-",
            report.updated_at.strftime("%Y-%m-%d %H:%M") if report.updated_at else "-",
        ])

    # 8️⃣ Format all cells (skip merged title row)
    for column_cells in ws.iter_cols(min_row=2, max_row=ws.max_row):
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border_style
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 3, 20)

    # 9️⃣ Prepare response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"Strategic_Reports_{cycle.slug}.xlsx"
    wb.save(response)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

#
from django.db.models import Count, Avg, F
from django.db.models.functions import TruncMonth
from django.utils import timezone
from datetime import datetime
from collections import defaultdict
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from plotly.offline import plot


@login_required
def strategic_report_chart(request):
    """Complete strategic dashboard using Django aggregates and Plotly."""

    # Color palette for charts
    COLORS = ['#4e73df', '#1cc88a', '#36b9cc', '#f6c23e', '#e74a3b', '#858796',
              '#5a5c69', '#6f42c1', '#e83e8c', '#fd7e14', '#20c997', '#6610f2']

    # Get filters
    cycle_filter = request.GET.get("strategic_cycle", "all")
    body_filter = request.GET.get("responsible_body", "all")

    # Base queryset
    reports = StrategicReport.objects.select_related(
        "action_plan__strategic_cycle",
        "action_plan__strategy_hierarchy"
    ).prefetch_related("action_plan__responsible_bodies").filter(
        organization_name=request.user.organization_name
    )

    # Apply filters
    if cycle_filter != "all":
        reports = reports.filter(action_plan__strategic_cycle_id=cycle_filter)
    if body_filter != "all":
        reports = reports.filter(action_plan__responsible_bodies__stakeholder_name=body_filter)

    # Get filter options
    strategic_cycles = StrategicCycle.objects.filter(
        organization_name=request.user.organization_name
    )
    responsible_bodies = list(reports.values_list(
        'action_plan__responsible_bodies__stakeholder_name', flat=True
    ).distinct().exclude(
        action_plan__responsible_bodies__stakeholder_name__isnull=True
    ).order_by('action_plan__responsible_bodies__stakeholder_name'))

    # 1. CORE METRICS & COUNTS
    total_reports = reports.count()

    status_counts = list(reports.values('status').annotate(
        count=Count('id')
    ).order_by('-count'))

    # Counts by cycle and body
    cycle_counts = list(reports.values(
        'action_plan__strategic_cycle__name'
    ).annotate(count=Count('id')).order_by('-count')) if cycle_filter == "all" else []

    body_counts = list(reports.values(
        'action_plan__responsible_bodies__stakeholder_name'
    ).annotate(count=Count('id')).filter(
        action_plan__responsible_bodies__stakeholder_name__isnull=False
    ).order_by('-count')) if body_filter == "all" else []

    overall_metrics = reports.aggregate(
        achievement=Avg('percent_achieved'),
        weighted_score=Avg('weighted_score')
    )

    # 2. MONTHLY PERFORMANCE
    monthly_data = list(reports.filter(
        action_plan__strategic_cycle__end_date__isnull=False
    ).annotate(
        month=TruncMonth('action_plan__strategic_cycle__end_date')
    ).values('month').annotate(
        achievement=Avg('percent_achieved'),
        weighted_score=Avg('weighted_score'),
        report_count=Count('id')
    ).order_by('month'))

    monthly_metrics = []
    for metric in monthly_data:
        if metric['month']:
            monthly_metrics.append({
                'month': metric['month'].strftime("%b %Y"),
                'full_date': metric['month'].strftime("%B %d, %Y"),
                'achievement': metric['achievement'] or 0,
                'weighted_score': metric['weighted_score'] or 0,
                'report_count': metric['report_count']
            })

    # 3. STAKEHOLDER PERFORMANCE
    stakeholder_monthly_data = list(reports.filter(
        action_plan__strategic_cycle__end_date__isnull=False,
        action_plan__responsible_bodies__isnull=False
    ).annotate(
        month=TruncMonth('action_plan__strategic_cycle__end_date'),
        stakeholder=F('action_plan__responsible_bodies__stakeholder_name')
    ).values('month', 'stakeholder').annotate(
        achievement=Avg('percent_achieved'),
        weighted_score=Avg('weighted_score'),
        report_count=Count('id')
    ).order_by('month', 'stakeholder'))

    # Organize stakeholder data
    stakeholder_by_month = defaultdict(dict)
    stakeholder_performance = defaultdict(lambda: {'achievement': [], 'count': 0, 'monthly': {}})

    for item in stakeholder_monthly_data:
        month_key = item['month'].strftime("%b %Y") if item['month'] else "Unknown"
        stakeholder = item['stakeholder']

        # For heatmap and trends
        stakeholder_by_month[month_key][stakeholder] = item['achievement'] or 0

        # For performance tracking
        stakeholder_performance[stakeholder]['achievement'].append(item['achievement'] or 0)
        stakeholder_performance[stakeholder]['count'] += item['report_count']
        stakeholder_performance[stakeholder]['monthly'][month_key] = item['achievement'] or 0

    # Calculate averages
    for stakeholder in stakeholder_performance:
        achievements = stakeholder_performance[stakeholder]['achievement']
        stakeholder_performance[stakeholder]['avg_achievement'] = sum(achievements) / len(
            achievements) if achievements else 0

    # REMOVED: Top stakeholders limit - show all stakeholders
    all_stakeholders = sorted(
        [(s, data) for s, data in stakeholder_performance.items()],
        key=lambda x: x[1]['count'], reverse=True
    )

    # 4. OBJECTIVE & KPI PERFORMANCE
    # Monthly objective data
    objective_monthly_data = list(reports.filter(
        action_plan__strategic_cycle__end_date__isnull=False,
        action_plan__strategy_hierarchy__objective__isnull=False
    ).annotate(
        month=TruncMonth('action_plan__strategic_cycle__end_date'),
        objective=F('action_plan__strategy_hierarchy__objective')
    ).values('month', 'objective').annotate(
        achievement=Avg('percent_achieved'),
        weighted_score=Avg('weighted_score'),
        report_count=Count('id')
    ).order_by('month', 'objective'))

    # Organize objective data
    objective_by_month = defaultdict(dict)
    objective_performance = defaultdict(
        lambda: {'achievement': [], 'weighted': [], 'count': 0, 'monthly_achievement': {}, 'monthly_weighted': {}})

    for item in objective_monthly_data:
        month_key = item['month'].strftime("%b %Y") if item['month'] else "Unknown"
        objective = item['objective']

        objective_by_month[month_key][objective] = item['achievement'] or 0

        objective_performance[objective]['achievement'].append(item['achievement'] or 0)
        objective_performance[objective]['weighted'].append(item['weighted_score'] or 0)
        objective_performance[objective]['count'] += item['report_count']
        objective_performance[objective]['monthly_achievement'][month_key] = item['achievement'] or 0
        objective_performance[objective]['monthly_weighted'][month_key] = item['weighted_score'] or 0

    # Calculate averages
    for objective in objective_performance:
        achievements = objective_performance[objective]['achievement']
        weighteds = objective_performance[objective]['weighted']
        objective_performance[objective]['avg_achievement'] = sum(achievements) / len(
            achievements) if achievements else 0
        objective_performance[objective]['avg_weighted'] = sum(weighteds) / len(weighteds) if weighteds else 0

    # REMOVED: Top objectives limit - show all objectives
    all_objectives = sorted(
        [(obj, data) for obj, data in objective_performance.items()],
        key=lambda x: x[1]['count'], reverse=True
    )

    # Monthly KPI data
    kpi_monthly_data = list(reports.filter(
        action_plan__strategic_cycle__end_date__isnull=False,
        action_plan__strategy_hierarchy__kpi__isnull=False
    ).annotate(
        month=TruncMonth('action_plan__strategic_cycle__end_date'),
        kpi=F('action_plan__strategy_hierarchy__kpi')
    ).values('month', 'kpi').annotate(
        achievement=Avg('percent_achieved'),
        weighted_score=Avg('weighted_score'),
        report_count=Count('id')
    ).order_by('month', 'kpi'))

    # Organize KPI data
    kpi_by_month = defaultdict(dict)
    kpi_performance = defaultdict(
        lambda: {'achievement': [], 'weighted': [], 'count': 0, 'monthly_achievement': {}, 'monthly_weighted': {}})

    for item in kpi_monthly_data:
        month_key = item['month'].strftime("%b %Y") if item['month'] else "Unknown"
        kpi = item['kpi']

        kpi_by_month[month_key][kpi] = item['achievement'] or 0

        kpi_performance[kpi]['achievement'].append(item['achievement'] or 0)
        kpi_performance[kpi]['weighted'].append(item['weighted_score'] or 0)
        kpi_performance[kpi]['count'] += item['report_count']
        kpi_performance[kpi]['monthly_achievement'][month_key] = item['achievement'] or 0
        kpi_performance[kpi]['monthly_weighted'][month_key] = item['weighted_score'] or 0

    # Calculate averages
    for kpi in kpi_performance:
        achievements = kpi_performance[kpi]['achievement']
        weighteds = kpi_performance[kpi]['weighted']
        kpi_performance[kpi]['avg_achievement'] = sum(achievements) / len(achievements) if achievements else 0
        kpi_performance[kpi]['avg_weighted'] = sum(weighteds) / len(weighteds) if weighteds else 0

    # REMOVED: Top KPIs limit - show all KPIs
    all_kpis = sorted(
        [(kpi, data) for kpi, data in kpi_performance.items()],
        key=lambda x: x[1]['count'], reverse=True
    )

    # 5. FILTER-SPECIFIC DATA
    objectives_by_body = []
    kpis_by_body = []

    if body_filter != "all":
        objectives_by_body = list(reports.filter(
            action_plan__responsible_bodies__stakeholder_name=body_filter,
            action_plan__strategy_hierarchy__objective__isnull=False
        ).values(
            'action_plan__strategy_hierarchy__objective'
        ).annotate(
            achievement=Avg('percent_achieved'),
            weighted_score=Avg('weighted_score'),
            report_count=Count('id')
        ).order_by('-report_count'))

        kpis_by_body = list(reports.filter(
            action_plan__responsible_bodies__stakeholder_name=body_filter,
            action_plan__strategy_hierarchy__kpi__isnull=False
        ).values(
            'action_plan__strategy_hierarchy__kpi'
        ).annotate(
            achievement=Avg('percent_achieved'),
            weighted_score=Avg('weighted_score'),
            report_count=Count('id')
        ).order_by('-report_count'))

    # 6. CREATE ALL CHARTS

    # Chart 1: Status Distribution
    if status_counts:
        status_map = {
            "pending": {"name": "Pending", "color": "#FFCE56"},
            "in_progress": {"name": "In Progress", "color": "#36A2EB"},
            "completed": {"name": "Completed", "color": "#4BC0C0"},
            "on_hold": {"name": "On Hold", "color": "#FF6384"},
            "cancelled": {"name": "Cancelled", "color": "#9966FF"},
        }

        labels = [status_map.get(s['status'], {}).get('name', s['status']) for s in status_counts]
        values = [s['count'] for s in status_counts]
        colors = [status_map.get(s['status'], {}).get('color', COLORS[0]) for s in status_counts]

        fig = go.Figure(data=[go.Pie(
            labels=labels, values=values, hole=0.5,
            marker=dict(colors=colors),
            textinfo='percent+label',
            hovertemplate='<b>%{label}</b><br>Count: %{value}<br>Percentage: %{percent}<extra></extra>'
        )])
        fig.update_layout(
            title_text='Report Status Distribution',
            height=400,
            annotations=[dict(text='Status', x=0.5, y=0.5, font_size=16, showarrow=False)]
        )
        status_plot = plot(fig, output_type='div', config={'displayModeBar': False})
    else:
        status_plot = "<div class='no-data'>No status data available</div>"

    # Chart 2: Monthly Performance Overview
    if monthly_metrics:
        months = [m['month'] for m in monthly_metrics]
        achievement = [m['achievement'] for m in monthly_metrics]
        weighted = [m['weighted_score'] for m in monthly_metrics]

        fig = make_subplots(
            rows=1, cols=2,
            subplot_titles=('Average % Achieved by Month', 'Average Weighted Score by Month')
        )

        fig.add_trace(go.Bar(
            x=months, y=achievement, name='% Achieved',
            marker_color=COLORS[0],
            hovertemplate='<b>%{x}</b><br>% Achieved: %{y:.2f}%<extra></extra>'
        ), 1, 1)

        fig.add_trace(go.Bar(
            x=months, y=weighted, name='Weighted Score',
            marker_color=COLORS[1],
            hovertemplate='<b>%{x}</b><br>Weighted Score: %{y:.2f}<extra></extra>'
        ), 1, 2)

        fig.update_layout(height=400, showlegend=False)
        fig.update_xaxes(tickangle=-45)
        overview_plot = plot(fig, output_type='div', config={'displayModeBar': False})
    else:
        overview_plot = "<div class='no-data'>No monthly data available</div>"

    # Chart 3: Stakeholder Heatmap
    if stakeholder_by_month and all_stakeholders:
        months = sorted(stakeholder_by_month.keys(),
                        key=lambda x: datetime.strptime(x, "%b %Y") if x != "Unknown" else datetime.min)
        # REMOVED: Top stakeholder limit - use all stakeholders
        stakeholder_names = [s[0] for s in all_stakeholders]

        z_data = []
        for stakeholder in stakeholder_names:
            row = []
            for month in months:
                row.append(stakeholder_by_month[month].get(stakeholder, 0))
            z_data.append(row)

        fig = go.Figure(data=go.Heatmap(
            z=z_data, x=months, y=stakeholder_names,
            colorscale='Viridis', hoverongaps=False,
            hovertemplate='<b>%{y}</b><br>Month: %{x}<br>% Achieved: %{z:.2f}%<extra></extra>'
        ))
        fig.update_layout(
            title='Department Performance Heatmap',
            xaxis_title='Month', yaxis_title='Departments',
            height=max(600, len(stakeholder_names) * 30),  # Dynamic height based on number of stakeholders
            xaxis=dict(tickangle=-45)
        )
        heatmap_plot = plot(fig, output_type='div', config={'displayModeBar': False})
    else:
        heatmap_plot = "<div class='no-data'>No stakeholder data available</div>"

    # Chart 4: Stakeholder Trends
    if stakeholder_by_month and all_stakeholders:
        months = sorted(stakeholder_by_month.keys(),
                        key=lambda x: datetime.strptime(x, "%b %Y") if x != "Unknown" else datetime.min)
        fig = go.Figure()

        # REMOVED: Limit of 8 stakeholders - show all stakeholders
        for i, (stakeholder, data) in enumerate(all_stakeholders):
            values = [data['monthly'].get(month, 0) for month in months]
            fig.add_trace(go.Scatter(
                x=months, y=values, mode='lines+markers',
                name=stakeholder[:20] + '...' if len(stakeholder) > 20 else stakeholder,
                line=dict(color=COLORS[i % len(COLORS)], width=2),
                marker=dict(size=4),
                hovertemplate=f'<b>{stakeholder}</b><br>Month: %{{x}}<br>% Achieved: %{{y:.2f}}%<extra></extra>'
            ))

        fig.update_layout(
            title='Department Performance Trends',
            xaxis_title='Month', yaxis_title='% Achieved',
            height=500, xaxis=dict(tickangle=-45),
            showlegend=True
        )
        trends_plot = plot(fig, output_type='div', config={'displayModeBar': False})
    else:
        trends_plot = "<div class='no-data'>No trend data available</div>"

    # Chart 5: Performance Distribution Boxplot
    if stakeholder_by_month:
        months = sorted(stakeholder_by_month.keys(),
                        key=lambda x: datetime.strptime(x, "%b %Y") if x != "Unknown" else datetime.min)

        monthly_performance_data = []
        for month in months:
            performances = list(stakeholder_by_month[month].values())
            monthly_performance_data.append(performances)

        fig = go.Figure()
        for i, month in enumerate(months):
            fig.add_trace(go.Box(
                y=monthly_performance_data[i], name=month,
                boxpoints='outliers', marker_color=COLORS[i % len(COLORS)]
            ))

        fig.update_layout(
            title='Performance Distribution by Month',
            xaxis_title='Month', yaxis_title='% Achieved',
            height=500, showlegend=False,
            xaxis=dict(tickangle=-45)
        )
        boxplot_plot = plot(fig, output_type='div', config={'displayModeBar': False})
    else:
        boxplot_plot = "<div class='no-data'>No performance data available</div>"

    # Chart 6: Objective Achievement
    if all_objectives and objective_by_month:
        months = sorted(objective_by_month.keys(),
                        key=lambda x: datetime.strptime(x, "%b %Y") if x != "Unknown" else datetime.min)
        fig = go.Figure()

        # REMOVED: Limit of top objectives - show all objectives
        for i, (objective, data) in enumerate(all_objectives):
            values = [data['monthly_achievement'].get(month, 0) for month in months]
            fig.add_trace(go.Bar(
                name=objective[:30] + '...' if len(objective) > 30 else objective,
                x=months, y=values, marker_color=COLORS[i % len(COLORS)]
            ))

        fig.update_layout(
            title='Objectives - % Achieved by Month',
            xaxis_title='Month', yaxis_title='% Achieved',
            height=500, barmode='group',
            xaxis=dict(tickangle=-45)
        )
        objective_achievement_plot = plot(fig, output_type='div', config={'displayModeBar': False})
    else:
        objective_achievement_plot = "<div class='no-data'>No objective data available</div>"

    # Chart 7: Objective Weighted Scores
    if all_objectives and objective_by_month:
        months = sorted(objective_by_month.keys(),
                        key=lambda x: datetime.strptime(x, "%b %Y") if x != "Unknown" else datetime.min)
        fig = go.Figure()

        # REMOVED: Limit of top objectives - show all objectives
        for i, (objective, data) in enumerate(all_objectives):
            values = [data['monthly_weighted'].get(month, 0) for month in months]
            fig.add_trace(go.Bar(
                name=objective[:30] + '...' if len(objective) > 30 else objective,
                x=months, y=values, marker_color=COLORS[i % len(COLORS)]
            ))

        fig.update_layout(
            title='Objectives - Weighted Score by Month',
            xaxis_title='Month', yaxis_title='Weighted Score',
            height=500, barmode='group',
            xaxis=dict(tickangle=-45)
        )
        objective_weighted_plot = plot(fig, output_type='div', config={'displayModeBar': False})
    else:
        objective_weighted_plot = "<div class='no-data'>No objective weighted data available</div>"

    # Chart 8: KPI Achievement
    if all_kpis and kpi_by_month:
        months = sorted(kpi_by_month.keys(),
                        key=lambda x: datetime.strptime(x, "%b %Y") if x != "Unknown" else datetime.min)
        fig = go.Figure()

        # REMOVED: Limit of top KPIs - show all KPIs
        for i, (kpi, data) in enumerate(all_kpis):
            values = [data['monthly_achievement'].get(month, 0) for month in months]
            fig.add_trace(go.Bar(
                name=kpi[:30] + '...' if len(kpi) > 30 else kpi,
                x=months, y=values, marker_color=COLORS[i % len(COLORS)]
            ))

        fig.update_layout(
            title='KPIs - % Achieved by Month',
            xaxis_title='Month', yaxis_title='% Achieved',
            height=500, barmode='group',
            xaxis=dict(tickangle=-45)
        )
        kpi_achievement_plot = plot(fig, output_type='div', config={'displayModeBar': False})
    else:
        kpi_achievement_plot = "<div class='no-data'>No KPI data available</div>"

    # Chart 9: KPI Weighted Scores
    if all_kpis and kpi_by_month:
        months = sorted(kpi_by_month.keys(),
                        key=lambda x: datetime.strptime(x, "%b %Y") if x != "Unknown" else datetime.min)
        fig = go.Figure()

        # REMOVED: Limit of top KPIs - show all KPIs
        for i, (kpi, data) in enumerate(all_kpis):
            values = [data['monthly_weighted'].get(month, 0) for month in months]
            fig.add_trace(go.Bar(
                name=kpi[:30] + '...' if len(kpi) > 30 else kpi,
                x=months, y=values, marker_color=COLORS[i % len(COLORS)]
            ))

        fig.update_layout(
            title='KPIs - Weighted Score by Month',
            xaxis_title='Month', yaxis_title='Weighted Score',
            height=500, barmode='group',
            xaxis=dict(tickangle=-45)
        )
        kpi_weighted_plot = plot(fig, output_type='div', config={'displayModeBar': False})
    else:
        kpi_weighted_plot = "<div class='no-data'>No KPI weighted data available</div>"

    # NEW: Stakeholder Performance Table Data
    # Enhanced stakeholder performance with weighted scores
    stakeholder_performance_with_weighted = defaultdict(lambda: {
        'achievement': [],
        'weighted_score': [],
        'count': 0,
        'monthly_achievement': {},
        'monthly_weighted': {}
    })

    for item in stakeholder_monthly_data:
        month_key = item['month'].strftime("%b %Y") if item['month'] else "Unknown"
        stakeholder = item['stakeholder']

        stakeholder_performance_with_weighted[stakeholder]['achievement'].append(item['achievement'] or 0)
        stakeholder_performance_with_weighted[stakeholder]['weighted_score'].append(item['weighted_score'] or 0)
        stakeholder_performance_with_weighted[stakeholder]['count'] += item['report_count']
        stakeholder_performance_with_weighted[stakeholder]['monthly_achievement'][month_key] = item['achievement'] or 0
        stakeholder_performance_with_weighted[stakeholder]['monthly_weighted'][month_key] = item['weighted_score'] or 0

    # Calculate averages for enhanced stakeholder data
    for stakeholder in stakeholder_performance_with_weighted:
        achievements = stakeholder_performance_with_weighted[stakeholder]['achievement']
        weighteds = stakeholder_performance_with_weighted[stakeholder]['weighted_score']
        stakeholder_performance_with_weighted[stakeholder]['avg_achievement'] = sum(achievements) / len(
            achievements) if achievements else 0
        stakeholder_performance_with_weighted[stakeholder]['avg_weighted'] = sum(weighteds) / len(
            weighteds) if weighteds else 0

    # Convert to list of tuples for template iteration
    stakeholder_performance_list = sorted(
        [(stakeholder, data) for stakeholder, data in stakeholder_performance_with_weighted.items()],
        key=lambda x: x[1]['count'], reverse=True
    )

    # NEW: Stakeholder Performance Comparison Chart - Show ALL stakeholders
    if stakeholder_performance_with_weighted:
        # Get ALL stakeholders for comparison
        all_stakeholders_comparison = stakeholder_performance_list

        stakeholders = [s[0] for s in all_stakeholders_comparison]
        avg_achievements = [s[1]['avg_achievement'] for s in all_stakeholders_comparison]
        avg_weighted = [s[1]['avg_weighted'] for s in all_stakeholders_comparison]

        fig = make_subplots(specs=[[{"secondary_y": True}]])

        fig.add_trace(
            go.Bar(
                x=stakeholders,
                y=avg_achievements,
                name="Avg % Achieved",
                marker_color=COLORS[0],
                hovertemplate='<b>%{x}</b><br>Avg Achievement: %{y:.1f}%<extra></extra>'
            ),
            secondary_y=False,
        )

        fig.add_trace(
            go.Scatter(
                x=stakeholders,
                y=avg_weighted,
                name="Avg Weighted Score",
                marker_color=COLORS[1],
                mode='lines+markers',
                hovertemplate='<b>%{x}</b><br>Avg Weighted: %{y:.2f}<extra></extra>'
            ),
            secondary_y=True,
        )

        fig.update_layout(
            title_text='Stakeholder Performance Comparison',
            xaxis_title='Departments',
            height=max(600, len(stakeholders) * 30),  # Dynamic height based on number of stakeholders
            xaxis=dict(tickangle=-45)
        )

        fig.update_yaxes(title_text="Average % Achieved", secondary_y=False)
        fig.update_yaxes(title_text="Average Weighted Score", secondary_y=True)

        stakeholder_comparison_plot = plot(fig, output_type='div', config={'displayModeBar': False})
    else:
        stakeholder_comparison_plot = "<div class='no-data'>No stakeholder comparison data available</div>"

    # Prepare context
    context = {
        # Core metrics
        'total_reports': total_reports,
        'status_counts': status_counts,
        'cycle_counts': cycle_counts,
        'body_counts': body_counts,
        'overall_metrics': {
            'achievement': overall_metrics['achievement'] or 0,
            'weighted_score': overall_metrics['weighted_score'] or 0,
            'percent_achieved': overall_metrics['achievement'] or 0,
        },
        'date_metrics': monthly_metrics,
        'body_metrics': list(reports.values(
            'action_plan__responsible_bodies__stakeholder_name'
        ).annotate(
            achievement=Avg('percent_achieved'),
            weighted_score=Avg('weighted_score'),
            report_count=Count('id')
        ).filter(
            action_plan__responsible_bodies__stakeholder_name__isnull=False
        ).order_by('-report_count')) if body_filter == "all" else [],

        # Data for tables - using all elements instead of top elements
        'top_objectives': all_objectives,
        'top_kpis': all_kpis,
        'top_stakeholders': all_stakeholders,

        # NEW: Enhanced stakeholder data with weighted scores for table display
        'stakeholder_performance_with_weighted': stakeholder_performance_list,  # Changed to list

        # Filter-specific data
        'objectives_by_body': objectives_by_body,
        'kpis_by_body': kpis_by_body,

        # All Charts
        'status_plot': status_plot,
        'overview_plot': overview_plot,
        'stakeholder_heatmap_plot': heatmap_plot,
        'stakeholder_line_plot': trends_plot,
        'performance_boxplot_plot': boxplot_plot,
        'objective_achievement_plot': objective_achievement_plot,
        'objective_weighted_plot': objective_weighted_plot,
        'kpi_achievement_plot': kpi_achievement_plot,
        'kpi_weighted_plot': kpi_weighted_plot,

        # NEW: Stakeholder comparison chart
        'stakeholder_comparison_plot': stakeholder_comparison_plot,

        # Filters
        'strategic_cycles': strategic_cycles,
        'responsible_bodies': responsible_bodies,
        'selected_cycle': cycle_filter,
        'selected_body': body_filter,
        'current_date': timezone.now(),

        # Filter states
        'cycle_filter_name': strategic_cycles.get(id=cycle_filter).name if cycle_filter != "all" else "All Cycles",
        'body_filter_name': body_filter if body_filter != "all" else "All Bodies",
    }

    return render(request, 'strategic_report/chart.html', context)
#
#
# from django.db.models import Count, Avg, F
# from django.db.models.functions import TruncMonth
# from django.utils import timezone
# from datetime import datetime
# from collections import defaultdict
# import plotly.graph_objects as go
# from plotly.subplots import make_subplots
# from plotly.offline import plot
#
# @login_required
# def strategic_report_chart(request):
#     """Complete strategic dashboard using Django aggregates and Plotly."""
#
#     # Color palette for charts
#     COLORS = ['#4e73df', '#1cc88a', '#36b9cc', '#f6c23e', '#e74a3b', '#858796',
#               '#5a5c69', '#6f42c1', '#e83e8c', '#fd7e14', '#20c997', '#6610f2']
#
#     # Get filters
#     cycle_filter = request.GET.get("strategic_cycle", "all")
#     body_filter = request.GET.get("responsible_body", "all")
#
#     # Base queryset
#     reports = StrategicReport.objects.select_related(
#         "action_plan__strategic_cycle",
#         "action_plan__strategy_hierarchy"
#     ).prefetch_related("action_plan__responsible_bodies").filter(
#         organization_name=request.user.organization_name
#     )
#
#     # Apply filters
#     if cycle_filter != "all":
#         reports = reports.filter(action_plan__strategic_cycle_id=cycle_filter)
#     if body_filter != "all":
#         reports = reports.filter(action_plan__responsible_bodies__stakeholder_name=body_filter)
#
#     # Get filter options
#     strategic_cycles = StrategicCycle.objects.filter(
#         organization_name=request.user.organization_name
#     )
#     responsible_bodies = list(reports.values_list(
#         'action_plan__responsible_bodies__stakeholder_name', flat=True
#     ).distinct().exclude(
#         action_plan__responsible_bodies__stakeholder_name__isnull=True
#     ).order_by('action_plan__responsible_bodies__stakeholder_name'))
#
#     # 1. CORE METRICS & COUNTS
#     total_reports = reports.count()
#
#     status_counts = list(reports.values('status').annotate(
#         count=Count('id')
#     ).order_by('-count'))
#
#     # Counts by cycle and body
#     cycle_counts = list(reports.values(
#         'action_plan__strategic_cycle__name'
#     ).annotate(count=Count('id')).order_by('-count')) if cycle_filter == "all" else []
#
#     body_counts = list(reports.values(
#         'action_plan__responsible_bodies__stakeholder_name'
#     ).annotate(count=Count('id')).filter(
#         action_plan__responsible_bodies__stakeholder_name__isnull=False
#     ).order_by('-count')) if body_filter == "all" else []
#
#     overall_metrics = reports.aggregate(
#         achievement=Avg('percent_achieved'),
#         weighted_score=Avg('weighted_score')
#     )
#
#     # 2. MONTHLY PERFORMANCE
#     monthly_data = list(reports.filter(
#         action_plan__strategic_cycle__end_date__isnull=False
#     ).annotate(
#         month=TruncMonth('action_plan__strategic_cycle__end_date')
#     ).values('month').annotate(
#         achievement=Avg('percent_achieved'),
#         weighted_score=Avg('weighted_score'),
#         report_count=Count('id')
#     ).order_by('month'))
#
#     monthly_metrics = []
#     for metric in monthly_data:
#         if metric['month']:
#             monthly_metrics.append({
#                 'month': metric['month'].strftime("%b %Y"),
#                 'full_date': metric['month'].strftime("%B %d, %Y"),
#                 'achievement': metric['achievement'] or 0,
#                 'weighted_score': metric['weighted_score'] or 0,
#                 'report_count': metric['report_count']
#             })
#
#     # 3. STAKEHOLDER PERFORMANCE
#     stakeholder_monthly_data = list(reports.filter(
#         action_plan__strategic_cycle__end_date__isnull=False,
#         action_plan__responsible_bodies__isnull=False
#     ).annotate(
#         month=TruncMonth('action_plan__strategic_cycle__end_date'),
#         stakeholder=F('action_plan__responsible_bodies__stakeholder_name')
#     ).values('month', 'stakeholder').annotate(
#         achievement=Avg('percent_achieved'),
#         weighted_score=Avg('weighted_score'),
#         report_count=Count('id')
#     ).order_by('month', 'stakeholder'))
#
#     # Organize stakeholder data
#     stakeholder_by_month = defaultdict(dict)
#     stakeholder_performance = defaultdict(lambda: {'achievement': [], 'count': 0, 'monthly': {}})
#
#     for item in stakeholder_monthly_data:
#         month_key = item['month'].strftime("%b %Y") if item['month'] else "Unknown"
#         stakeholder = item['stakeholder']
#
#         # For heatmap and trends
#         stakeholder_by_month[month_key][stakeholder] = item['achievement'] or 0
#
#         # For performance tracking
#         stakeholder_performance[stakeholder]['achievement'].append(item['achievement'] or 0)
#         stakeholder_performance[stakeholder]['count'] += item['report_count']
#         stakeholder_performance[stakeholder]['monthly'][month_key] = item['achievement'] or 0
#
#     # Calculate averages
#     for stakeholder in stakeholder_performance:
#         achievements = stakeholder_performance[stakeholder]['achievement']
#         stakeholder_performance[stakeholder]['avg_achievement'] = sum(achievements) / len(
#             achievements) if achievements else 0
#
#     # REMOVED: Top stakeholders limit - show all stakeholders
#     all_stakeholders = sorted(
#         [(s, data) for s, data in stakeholder_performance.items()],
#         key=lambda x: x[1]['count'], reverse=True
#     )
#
#     # 4. OBJECTIVE & KPI PERFORMANCE
#     # Monthly objective data
#     objective_monthly_data = list(reports.filter(
#         action_plan__strategic_cycle__end_date__isnull=False,
#         action_plan__strategy_hierarchy__objective__isnull=False
#     ).annotate(
#         month=TruncMonth('action_plan__strategic_cycle__end_date'),
#         objective=F('action_plan__strategy_hierarchy__objective')
#     ).values('month', 'objective').annotate(
#         achievement=Avg('percent_achieved'),
#         weighted_score=Avg('weighted_score'),
#         report_count=Count('id')
#     ).order_by('month', 'objective'))
#
#     # Organize objective data
#     objective_by_month = defaultdict(dict)
#     objective_performance = defaultdict(
#         lambda: {'achievement': [], 'weighted': [], 'count': 0, 'monthly_achievement': {}, 'monthly_weighted': {}})
#
#     for item in objective_monthly_data:
#         month_key = item['month'].strftime("%b %Y") if item['month'] else "Unknown"
#         objective = item['objective']
#
#         objective_by_month[month_key][objective] = item['achievement'] or 0
#
#         objective_performance[objective]['achievement'].append(item['achievement'] or 0)
#         objective_performance[objective]['weighted'].append(item['weighted_score'] or 0)
#         objective_performance[objective]['count'] += item['report_count']
#         objective_performance[objective]['monthly_achievement'][month_key] = item['achievement'] or 0
#         objective_performance[objective]['monthly_weighted'][month_key] = item['weighted_score'] or 0
#
#     # Calculate averages
#     for objective in objective_performance:
#         achievements = objective_performance[objective]['achievement']
#         weighteds = objective_performance[objective]['weighted']
#         objective_performance[objective]['avg_achievement'] = sum(achievements) / len(
#             achievements) if achievements else 0
#         objective_performance[objective]['avg_weighted'] = sum(weighteds) / len(weighteds) if weighteds else 0
#
#     # REMOVED: Top objectives limit - show all objectives
#     all_objectives = sorted(
#         [(obj, data) for obj, data in objective_performance.items()],
#         key=lambda x: x[1]['count'], reverse=True
#     )
#
#     # Monthly KPI data
#     kpi_monthly_data = list(reports.filter(
#         action_plan__strategic_cycle__end_date__isnull=False,
#         action_plan__strategy_hierarchy__kpi__isnull=False
#     ).annotate(
#         month=TruncMonth('action_plan__strategic_cycle__end_date'),
#         kpi=F('action_plan__strategy_hierarchy__kpi')
#     ).values('month', 'kpi').annotate(
#         achievement=Avg('percent_achieved'),
#         weighted_score=Avg('weighted_score'),
#         report_count=Count('id')
#     ).order_by('month', 'kpi'))
#
#     # Organize KPI data
#     kpi_by_month = defaultdict(dict)
#     kpi_performance = defaultdict(
#         lambda: {'achievement': [], 'weighted': [], 'count': 0, 'monthly_achievement': {}, 'monthly_weighted': {}})
#
#     for item in kpi_monthly_data:
#         month_key = item['month'].strftime("%b %Y") if item['month'] else "Unknown"
#         kpi = item['kpi']
#
#         kpi_by_month[month_key][kpi] = item['achievement'] or 0
#
#         kpi_performance[kpi]['achievement'].append(item['achievement'] or 0)
#         kpi_performance[kpi]['weighted'].append(item['weighted_score'] or 0)
#         kpi_performance[kpi]['count'] += item['report_count']
#         kpi_performance[kpi]['monthly_achievement'][month_key] = item['achievement'] or 0
#         kpi_performance[kpi]['monthly_weighted'][month_key] = item['weighted_score'] or 0
#
#     # Calculate averages
#     for kpi in kpi_performance:
#         achievements = kpi_performance[kpi]['achievement']
#         weighteds = kpi_performance[kpi]['weighted']
#         kpi_performance[kpi]['avg_achievement'] = sum(achievements) / len(achievements) if achievements else 0
#         kpi_performance[kpi]['avg_weighted'] = sum(weighteds) / len(weighteds) if weighteds else 0
#
#     # REMOVED: Top KPIs limit - show all KPIs
#     all_kpis = sorted(
#         [(kpi, data) for kpi, data in kpi_performance.items()],
#         key=lambda x: x[1]['count'], reverse=True
#     )
#
#     # 5. FILTER-SPECIFIC DATA
#     objectives_by_body = []
#     kpis_by_body = []
#
#     if body_filter != "all":
#         objectives_by_body = list(reports.filter(
#             action_plan__responsible_bodies__stakeholder_name=body_filter,
#             action_plan__strategy_hierarchy__objective__isnull=False
#         ).values(
#             'action_plan__strategy_hierarchy__objective'
#         ).annotate(
#             achievement=Avg('percent_achieved'),
#             weighted_score=Avg('weighted_score'),
#             report_count=Count('id')
#         ).order_by('-report_count'))
#
#         kpis_by_body = list(reports.filter(
#             action_plan__responsible_bodies__stakeholder_name=body_filter,
#             action_plan__strategy_hierarchy__kpi__isnull=False
#         ).values(
#             'action_plan__strategy_hierarchy__kpi'
#         ).annotate(
#             achievement=Avg('percent_achieved'),
#             weighted_score=Avg('weighted_score'),
#             report_count=Count('id')
#         ).order_by('-report_count'))
#
#     # 6. CREATE ALL CHARTS
#
#     # Chart 1: Status Distribution
#     if status_counts:
#         status_map = {
#             "pending": {"name": "Pending", "color": "#FFCE56"},
#             "in_progress": {"name": "In Progress", "color": "#36A2EB"},
#             "completed": {"name": "Completed", "color": "#4BC0C0"},
#             "on_hold": {"name": "On Hold", "color": "#FF6384"},
#             "cancelled": {"name": "Cancelled", "color": "#9966FF"},
#         }
#
#         labels = [status_map.get(s['status'], {}).get('name', s['status']) for s in status_counts]
#         values = [s['count'] for s in status_counts]
#         colors = [status_map.get(s['status'], {}).get('color', COLORS[0]) for s in status_counts]
#
#         fig = go.Figure(data=[go.Pie(
#             labels=labels, values=values, hole=0.5,
#             marker=dict(colors=colors),
#             textinfo='percent+label',
#             hovertemplate='<b>%{label}</b><br>Count: %{value}<br>Percentage: %{percent}<extra></extra>'
#         )])
#         fig.update_layout(
#             title_text='Report Status Distribution',
#             height=400,
#             annotations=[dict(text='Status', x=0.5, y=0.5, font_size=16, showarrow=False)]
#         )
#         status_plot = plot(fig, output_type='div', config={'displayModeBar': False})
#     else:
#         status_plot = "<div class='no-data'>No status data available</div>"
#
#     # Chart 2: Monthly Performance Overview
#     if monthly_metrics:
#         months = [m['month'] for m in monthly_metrics]
#         achievement = [m['achievement'] for m in monthly_metrics]
#         weighted = [m['weighted_score'] for m in monthly_metrics]
#
#         fig = make_subplots(
#             rows=1, cols=2,
#             subplot_titles=('Average % Achieved by Month', 'Average Weighted Score by Month')
#         )
#
#         fig.add_trace(go.Bar(
#             x=months, y=achievement, name='% Achieved',
#             marker_color=COLORS[0],
#             hovertemplate='<b>%{x}</b><br>% Achieved: %{y:.2f}%<extra></extra>'
#         ), 1, 1)
#
#         fig.add_trace(go.Bar(
#             x=months, y=weighted, name='Weighted Score',
#             marker_color=COLORS[1],
#             hovertemplate='<b>%{x}</b><br>Weighted Score: %{y:.2f}<extra></extra>'
#         ), 1, 2)
#
#         fig.update_layout(height=400, showlegend=False)
#         fig.update_xaxes(tickangle=-45)
#         overview_plot = plot(fig, output_type='div', config={'displayModeBar': False})
#     else:
#         overview_plot = "<div class='no-data'>No monthly data available</div>"
#
#     # Chart 3: Stakeholder Heatmap
#     if stakeholder_by_month and all_stakeholders:
#         months = sorted(stakeholder_by_month.keys(),
#                         key=lambda x: datetime.strptime(x, "%b %Y") if x != "Unknown" else datetime.min)
#         # REMOVED: Top stakeholder limit - use all stakeholders
#         stakeholder_names = [s[0] for s in all_stakeholders]
#
#         z_data = []
#         for stakeholder in stakeholder_names:
#             row = []
#             for month in months:
#                 row.append(stakeholder_by_month[month].get(stakeholder, 0))
#             z_data.append(row)
#
#         fig = go.Figure(data=go.Heatmap(
#             z=z_data, x=months, y=stakeholder_names,
#             colorscale='Viridis', hoverongaps=False,
#             hovertemplate='<b>%{y}</b><br>Month: %{x}<br>% Achieved: %{z:.2f}%<extra></extra>'
#         ))
#         fig.update_layout(
#             title='Department Performance Heatmap',
#             xaxis_title='Month', yaxis_title='Departments',
#             height=max(600, len(stakeholder_names) * 30),  # Dynamic height based on number of stakeholders
#             xaxis=dict(tickangle=-45)
#         )
#         heatmap_plot = plot(fig, output_type='div', config={'displayModeBar': False})
#     else:
#         heatmap_plot = "<div class='no-data'>No stakeholder data available</div>"
#
#     # Chart 4: Stakeholder Trends
#     if stakeholder_by_month and all_stakeholders:
#         months = sorted(stakeholder_by_month.keys(),
#                         key=lambda x: datetime.strptime(x, "%b %Y") if x != "Unknown" else datetime.min)
#         fig = go.Figure()
#
#         # REMOVED: Limit of 8 stakeholders - show all stakeholders
#         for i, (stakeholder, data) in enumerate(all_stakeholders):
#             values = [data['monthly'].get(month, 0) for month in months]
#             fig.add_trace(go.Scatter(
#                 x=months, y=values, mode='lines+markers',
#                 name=stakeholder[:20] + '...' if len(stakeholder) > 20 else stakeholder,
#                 line=dict(color=COLORS[i % len(COLORS)], width=2),
#                 marker=dict(size=4),
#                 hovertemplate=f'<b>{stakeholder}</b><br>Month: %{{x}}<br>% Achieved: %{{y:.2f}}%<extra></extra>'
#             ))
#
#         fig.update_layout(
#             title='Department Performance Trends',
#             xaxis_title='Month', yaxis_title='% Achieved',
#             height=500, xaxis=dict(tickangle=-45),
#             showlegend=True
#         )
#         trends_plot = plot(fig, output_type='div', config={'displayModeBar': False})
#     else:
#         trends_plot = "<div class='no-data'>No trend data available</div>"
#
#     # Chart 5: Performance Distribution Boxplot
#     if stakeholder_by_month:
#         months = sorted(stakeholder_by_month.keys(),
#                         key=lambda x: datetime.strptime(x, "%b %Y") if x != "Unknown" else datetime.min)
#
#         monthly_performance_data = []
#         for month in months:
#             performances = list(stakeholder_by_month[month].values())
#             monthly_performance_data.append(performances)
#
#         fig = go.Figure()
#         for i, month in enumerate(months):
#             fig.add_trace(go.Box(
#                 y=monthly_performance_data[i], name=month,
#                 boxpoints='outliers', marker_color=COLORS[i % len(COLORS)]
#             ))
#
#         fig.update_layout(
#             title='Performance Distribution by Month',
#             xaxis_title='Month', yaxis_title='% Achieved',
#             height=500, showlegend=False,
#             xaxis=dict(tickangle=-45)
#         )
#         boxplot_plot = plot(fig, output_type='div', config={'displayModeBar': False})
#     else:
#         boxplot_plot = "<div class='no-data'>No performance data available</div>"
#
#     # Chart 6: Objective Achievement
#     if all_objectives and objective_by_month:
#         months = sorted(objective_by_month.keys(),
#                         key=lambda x: datetime.strptime(x, "%b %Y") if x != "Unknown" else datetime.min)
#         fig = go.Figure()
#
#         # REMOVED: Limit of top objectives - show all objectives
#         for i, (objective, data) in enumerate(all_objectives):
#             values = [data['monthly_achievement'].get(month, 0) for month in months]
#             fig.add_trace(go.Bar(
#                 name=objective[:30] + '...' if len(objective) > 30 else objective,
#                 x=months, y=values, marker_color=COLORS[i % len(COLORS)]
#             ))
#
#         fig.update_layout(
#             title='Objectives - % Achieved by Month',
#             xaxis_title='Month', yaxis_title='% Achieved',
#             height=500, barmode='group',
#             xaxis=dict(tickangle=-45)
#         )
#         objective_achievement_plot = plot(fig, output_type='div', config={'displayModeBar': False})
#     else:
#         objective_achievement_plot = "<div class='no-data'>No objective data available</div>"
#
#     # Chart 7: Objective Weighted Scores
#     if all_objectives and objective_by_month:
#         months = sorted(objective_by_month.keys(),
#                         key=lambda x: datetime.strptime(x, "%b %Y") if x != "Unknown" else datetime.min)
#         fig = go.Figure()
#
#         # REMOVED: Limit of top objectives - show all objectives
#         for i, (objective, data) in enumerate(all_objectives):
#             values = [data['monthly_weighted'].get(month, 0) for month in months]
#             fig.add_trace(go.Bar(
#                 name=objective[:30] + '...' if len(objective) > 30 else objective,
#                 x=months, y=values, marker_color=COLORS[i % len(COLORS)]
#             ))
#
#         fig.update_layout(
#             title='Objectives - Weighted Score by Month',
#             xaxis_title='Month', yaxis_title='Weighted Score',
#             height=500, barmode='group',
#             xaxis=dict(tickangle=-45)
#         )
#         objective_weighted_plot = plot(fig, output_type='div', config={'displayModeBar': False})
#     else:
#         objective_weighted_plot = "<div class='no-data'>No objective weighted data available</div>"
#
#     # Chart 8: KPI Achievement
#     if all_kpis and kpi_by_month:
#         months = sorted(kpi_by_month.keys(),
#                         key=lambda x: datetime.strptime(x, "%b %Y") if x != "Unknown" else datetime.min)
#         fig = go.Figure()
#
#         # REMOVED: Limit of top KPIs - show all KPIs
#         for i, (kpi, data) in enumerate(all_kpis):
#             values = [data['monthly_achievement'].get(month, 0) for month in months]
#             fig.add_trace(go.Bar(
#                 name=kpi[:30] + '...' if len(kpi) > 30 else kpi,
#                 x=months, y=values, marker_color=COLORS[i % len(COLORS)]
#             ))
#
#         fig.update_layout(
#             title='KPIs - % Achieved by Month',
#             xaxis_title='Month', yaxis_title='% Achieved',
#             height=500, barmode='group',
#             xaxis=dict(tickangle=-45)
#         )
#         kpi_achievement_plot = plot(fig, output_type='div', config={'displayModeBar': False})
#     else:
#         kpi_achievement_plot = "<div class='no-data'>No KPI data available</div>"
#
#     # Chart 9: KPI Weighted Scores
#     if all_kpis and kpi_by_month:
#         months = sorted(kpi_by_month.keys(),
#                         key=lambda x: datetime.strptime(x, "%b %Y") if x != "Unknown" else datetime.min)
#         fig = go.Figure()
#
#         # REMOVED: Limit of top KPIs - show all KPIs
#         for i, (kpi, data) in enumerate(all_kpis):
#             values = [data['monthly_weighted'].get(month, 0) for month in months]
#             fig.add_trace(go.Bar(
#                 name=kpi[:30] + '...' if len(kpi) > 30 else kpi,
#                 x=months, y=values, marker_color=COLORS[i % len(COLORS)]
#             ))
#
#         fig.update_layout(
#             title='KPIs - Weighted Score by Month',
#             xaxis_title='Month', yaxis_title='Weighted Score',
#             height=500, barmode='group',
#             xaxis=dict(tickangle=-45)
#         )
#         kpi_weighted_plot = plot(fig, output_type='div', config={'displayModeBar': False})
#     else:
#         kpi_weighted_plot = "<div class='no-data'>No KPI weighted data available</div>"
#
#     # Prepare context
#     context = {
#         # Core metrics
#         'total_reports': total_reports,
#         'status_counts': status_counts,
#         'cycle_counts': cycle_counts,
#         'body_counts': body_counts,
#         'overall_metrics': {
#             'achievement': overall_metrics['achievement'] or 0,
#             'weighted_score': overall_metrics['weighted_score'] or 0,
#             'percent_achieved': overall_metrics['achievement'] or 0,
#         },
#         'date_metrics': monthly_metrics,
#         'body_metrics': list(reports.values(
#             'action_plan__responsible_bodies__stakeholder_name'
#         ).annotate(
#             achievement=Avg('percent_achieved'),
#             weighted_score=Avg('weighted_score'),
#             report_count=Count('id')
#         ).filter(
#             action_plan__responsible_bodies__stakeholder_name__isnull=False
#         ).order_by('-report_count')) if body_filter == "all" else [],
#
#         # Data for tables - using all elements instead of top elements
#         'top_objectives': all_objectives,
#         'top_kpis': all_kpis,
#         'top_stakeholders': all_stakeholders,
#
#         # Filter-specific data
#         'objectives_by_body': objectives_by_body,
#         'kpis_by_body': kpis_by_body,
#
#         # All Charts
#         'status_plot': status_plot,
#         'overview_plot': overview_plot,
#         'stakeholder_heatmap_plot': heatmap_plot,
#         'stakeholder_line_plot': trends_plot,
#         'performance_boxplot_plot': boxplot_plot,
#         'objective_achievement_plot': objective_achievement_plot,
#         'objective_weighted_plot': objective_weighted_plot,
#         'kpi_achievement_plot': kpi_achievement_plot,
#         'kpi_weighted_plot': kpi_weighted_plot,
#
#         # Filters
#         'strategic_cycles': strategic_cycles,
#         'responsible_bodies': responsible_bodies,
#         'selected_cycle': cycle_filter,
#         'selected_body': body_filter,
#         'current_date': timezone.now(),
#
#         # Filter states
#         'cycle_filter_name': strategic_cycles.get(id=cycle_filter).name if cycle_filter != "all" else "All Cycles",
#         'body_filter_name': body_filter if body_filter != "all" else "All Bodies",
#     }
#
#     return render(request, 'strategic_report/chart.html', context)

