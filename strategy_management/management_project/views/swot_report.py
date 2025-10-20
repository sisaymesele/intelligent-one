from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
import calendar
from django.core.paginator import Paginator
from django.urls import reverse

from management_project.models import SwotReport, StrategicCycle, StrategicReport
from management_project.forms import SwotReportForm

from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

from django.db.models import Count, Q
import plotly.graph_objects as go


@login_required
def swot_report_by_cycle_list(request):
    """List distinct strategic cycles for the current organization with all info."""
    cycles_qs = StrategicCycle.objects.filter(
        organization_name=request.user.organization_name
    ).order_by('-start_date')

    # Build a list of dicts including calculated properties
    cycles = []
    for cycle in cycles_qs:
        cycles.append({
            'name': cycle.name,
            'time_horizon': cycle.time_horizon,
            'time_horizon_type': cycle.time_horizon_type,
            'start_date': cycle.start_date,
            'end_date': cycle.end_date,
            'slug': cycle.slug,
            'duration_days': (cycle.end_date - cycle.start_date).days if cycle.start_date and cycle.end_date else None,
            'start_month_name': calendar.month_name[cycle.start_date.month] if cycle.start_date else None,
            'start_quarter': (cycle.start_date.month - 1) // 3 + 1 if cycle.start_date else None,
            'start_year': cycle.start_date.year if cycle.start_date else None,
        })

    return render(request, 'swot_report/cycle_list.html', {
        'strategic_cycles': cycles
    })


# -------------------- SWOT LIST --
@login_required
def swot_report_list(request):
    query = request.GET.get("search", "").strip()
    cycle_slug = request.GET.get("cycle", "").strip()
    page_number = request.GET.get("page", 1)

    # Get all SWOT reports for the organization
    swots = SwotReport.objects.filter(
        organization_name=request.user.organization_name
    )

    # Filter by strategic cycle if provided
    if cycle_slug:
        try:
            strategic_cycle = StrategicCycle.objects.get(
                slug=cycle_slug,
                organization_name=request.user.organization_name
            )
            strategic_reports = StrategicReport.objects.filter(
                action_plan__strategic_cycle=strategic_cycle
            )
            swots = swots.filter(strategic_report_period__in=strategic_reports)
        except StrategicCycle.DoesNotExist:
            pass

    # Apply search filter
    if query:
        swots = swots.filter(
            Q(swot_type__icontains=query) |
            Q(swot_pillar__icontains=query) |
            Q(swot_factor__icontains=query) |
            Q(description__icontains=query) |
            Q(strategic_report_period__action_plan__strategic_cycle__name__icontains=query) |
            Q(strategic_report_period__action_plan__strategic_cycle__time_horizon__icontains=query) |
            Q(strategic_report_period__action_plan__strategic_cycle__time_horizon_type__icontains=query)
        )

    # Available strategic cycles for filter dropdown
    strategic_cycles = StrategicCycle.objects.filter(
        organization_name=request.user.organization_name
    ).order_by('-start_date')

    # Ordering
    swots = swots.order_by("swot_type", "priority", "-created_at")

    # Pagination
    paginator = Paginator(swots, 10)  # 10 items per page
    page_obj = paginator.get_page(page_number)

    return render(request, "swot_report/list.html", {
        "swots": page_obj,
        "page_obj": page_obj,
        "search_query": query,
        "strategic_cycles": strategic_cycles,
        "selected_cycle": cycle_slug,
    })
#
#
# -------------------- CREATE SWOT --------------------


@login_required
def create_swot_report(request):
    org = request.user.organization_name
    # Get the latest StrategicReport for this organization
    latest_report = StrategicReport.objects.filter(organization_name=org).order_by('-id').first()

    if request.method == 'POST':
        form = SwotReportForm(request.POST, request=request)
        if 'save' in request.POST and form.is_valid():
            swot_entry = form.save(commit=False)
            swot_entry.organization_name = org
            # Assign the selected strategic report or fallback to latest
            if not swot_entry.strategic_report_period:
                swot_entry.strategic_report_period = latest_report
            swot_entry.save()
            messages.success(request, "SWOT entry created successfully!")
            return redirect('swot_report_list')
        else:
            messages.error(request, "Error creating SWOT entry. Please check the form.")
    else:
        form = SwotReportForm(request=request)
        if latest_report:
            form.fields['strategic_report_period'].initial = latest_report

    return render(request, 'swot_report/form.html', {'form': form})

# -------------------- UPDATE SWOT --------------------
@login_required
def update_swot_report(request, pk):
    entry = get_object_or_404(
        SwotReport,
        pk=pk,
        organization_name=request.user.organization_name
    )

    if request.method == 'POST':
        form = SwotReportForm(request.POST, instance=entry, request=request)
        if 'save' in request.POST and form.is_valid():
            form.save()
            messages.success(request, "SWOT entry updated successfully!")
            return redirect('swot_report_list')
        else:
            messages.error(request, "Error updating SWOT entry. Please check the form.")
    else:
        form = SwotReportForm(instance=entry, request=request)

    return render(request, 'swot_report/form.html', {'form': form})


# -------------------- DELETE SWOT --------------------
@login_required
def delete_swot_report(request, pk):
    entry = get_object_or_404(
        SwotReport,
        pk=pk,
        organization_name=request.user.organization_name
    )

    if request.method == 'POST':
        entry.delete()
        messages.success(request, "SWOT entry deleted successfully!")
        return redirect('swot_report_list')

    return render(request, 'swot_report/delete_confirm.html', {'entry': entry})



@login_required
def swot_report_list(request):
    query = request.GET.get("search", "").strip()
    cycle_slug = request.GET.get("cycle", "").strip()
    page_number = request.GET.get("page", 1)
    export = request.GET.get("export", "").lower()

    # Get all SWOT reports for the organization
    swots = SwotReport.objects.filter(
        organization_name=request.user.organization_name
    )

    # Filter by strategic cycle if provided
    if cycle_slug:
        try:
            strategic_cycle = StrategicCycle.objects.get(
                slug=cycle_slug,
                organization_name=request.user.organization_name
            )
            strategic_reports = StrategicReport.objects.filter(
                action_plan__strategic_cycle=strategic_cycle
            )
            swots = swots.filter(strategic_report_period__in=strategic_reports)
        except StrategicCycle.DoesNotExist:
            pass

    # Apply search filter
    if query:
        swots = swots.filter(
            Q(swot_type__icontains=query) |
            Q(swot_pillar__icontains=query) |
            Q(swot_factor__icontains=query) |
            Q(description__icontains=query) |
            Q(strategic_report_period__action_plan__strategic_cycle__name__icontains=query) |
            Q(strategic_report_period__action_plan__strategic_cycle__time_horizon__icontains=query) |
            Q(strategic_report_period__action_plan__strategic_cycle__time_horizon_type__icontains=query)
        )

    # Ordering
    swots = swots.order_by("swot_type", "priority", "-created_at")

    # Excel Export
    if export == "excel":
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=swot_reports.xlsx'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SWOT Reports"

        # Headers
        headers = [
            "Strategic Cycle", "SWOT Type", "Pillar", "Factor",
            "Description", "Priority", "Impact", "Likelihood", "Created At"
        ]
        ws.append(headers)

        # Data rows
        for s in swots:
            ws.append([
                s.strategic_report_period.action_plan.strategic_cycle.name,
                s.swot_type,
                s.swot_pillar,
                s.swot_factor,
                s.description,
                s.priority,
                s.impact,
                s.likelihood or "",
                s.created_at.strftime("%Y-%m-%d"),
            ])

        # Optional: adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(response)
        return response

    # Pagination
    paginator = Paginator(swots, 10)  # 10 items per page
    page_obj = paginator.get_page(page_number)

    # Available strategic cycles for filter dropdown
    strategic_cycles = StrategicCycle.objects.filter(
        organization_name=request.user.organization_name
    ).order_by('-start_date')

    return render(request, "swot_report/list.html", {
        "swots": page_obj,
        "page_obj": page_obj,
        "search_query": query,
        "strategic_cycles": strategic_cycles,
        "selected_cycle": cycle_slug,
    })



@login_required
def swot_report_chart(request):
    qs = SwotReport.objects.filter(
        organization_name=request.user.organization_name
    )

    # Optional: filter by strategic cycle from GET parameter
    cycle_filter = request.GET.get("cycle")
    if cycle_filter:
        qs = qs.filter(
            strategic_report_period__action_plan__strategic_cycle__name=cycle_filter
        )

    # ------------------ SWOT TYPE DISTRIBUTION ------------------
    type_counts = qs.values('swot_type').annotate(count=Count('id'))
    type_labels = [t['swot_type'] for t in type_counts]
    type_values = [t['count'] for t in type_counts]

    fig_swot_type = go.Figure(data=[go.Bar(
        x=type_labels,
        y=type_values,
        text=type_values,
        textposition='auto',
        marker_color=['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728']
    )])
    fig_swot_type.update_layout(
        title=f'SWOT Distribution by Type ({cycle_filter if cycle_filter else "All Cycles"})',
        xaxis_title='SWOT Type',
        yaxis_title='Count',
        template='plotly_white',
        height=400
    )

    # ------------------ PRIORITY DISTRIBUTION ------------------
    priority_counts = qs.values('priority').annotate(count=Count('id'))
    priority_labels = [p['priority'] for p in priority_counts]
    priority_values = [p['count'] for p in priority_counts]

    fig_priority = go.Figure(data=[go.Bar(
        x=priority_labels,
        y=priority_values,
        text=priority_values,
        textposition='auto',
        marker_color=['#d62728', '#ff7f0e', '#ffbb78', '#98df8a', '#2ca02c']
    )])
    fig_priority.update_layout(
        title=f'SWOT Distribution by Priority ({cycle_filter if cycle_filter else "All Cycles"})',
        xaxis_title='Priority',
        yaxis_title='Count',
        template='plotly_white',
        height=400
    )

    # ------------------ IMPACT DISTRIBUTION ------------------
    impact_counts = qs.values('impact').annotate(count=Count('id'))
    impact_labels = [i['impact'] for i in impact_counts]
    impact_values = [i['count'] for i in impact_counts]

    fig_impact = go.Figure(data=[go.Bar(
        x=impact_labels,
        y=impact_values,
        text=impact_values,
        textposition='auto',
        marker_color=['#98df8a', '#2ca02c', '#ffbb78', '#ff7f0e', '#d62728']
    )])
    fig_impact.update_layout(
        title=f'SWOT Distribution by Impact ({cycle_filter if cycle_filter else "All Cycles"})',
        xaxis_title='Impact',
        yaxis_title='Count',
        template='plotly_white',
        height=400
    )

    # ------------------ LIKELIHOOD DISTRIBUTION ------------------
    likelihood_counts = qs.values('likelihood').annotate(count=Count('id'))
    likelihood_labels = [l['likelihood'] if l['likelihood'] else 'Unknown' for l in likelihood_counts]
    likelihood_values = [l['count'] for l in likelihood_counts]

    fig_likelihood = go.Figure(data=[go.Bar(
        x=likelihood_labels,
        y=likelihood_values,
        text=likelihood_values,
        textposition='auto',
        marker_color=['#17becf', '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#7f7f7f']
    )])
    fig_likelihood.update_layout(
        title=f'SWOT Distribution by Likelihood ({cycle_filter if cycle_filter else "All Cycles"})',
        xaxis_title='Likelihood',
        yaxis_title='Count',
        template='plotly_white',
        height=400
    )

    # ------------------ PILLAR ANALYSIS ------------------
    pillar_counts = qs.values('swot_pillar').annotate(count=Count('id')).order_by('-count')
    pillar_labels = [p['swot_pillar'] for p in pillar_counts]
    pillar_values = [p['count'] for p in pillar_counts]

    fig_pillar = go.Figure(data=[go.Bar(
        x=pillar_labels,
        y=pillar_values,
        text=pillar_values,
        textposition='auto',
        marker_color=['#636efa', '#ef553b', '#00cc96', '#ab63fa', '#ffa15a']
    )])
    fig_pillar.update_layout(
        title=f'SWOT Count per Pillar ({cycle_filter if cycle_filter else "All Cycles"})',
        xaxis_title='Pillar',
        yaxis_title='Count',
        template='plotly_white',
        height=400
    )

    # ------------------ SUMMARY DATA ------------------
    total_swot = qs.count()
    high_priority_count = qs.filter(priority__in=['High', 'Very High']).count()
    strengths_count = qs.filter(swot_type='Strength').count()
    weaknesses_count = qs.filter(swot_type='Weakness').count()
    opportunities_count = qs.filter(swot_type='Opportunity').count()
    threats_count = qs.filter(swot_type='Threat').count()

    summary_data = {
        'total_swot': total_swot,
        'high_priority_count': high_priority_count,
        'strengths_count': strengths_count,
        'weaknesses_count': weaknesses_count,
        'opportunities_count': opportunities_count,
        'threats_count': threats_count,
        'cycle': cycle_filter if cycle_filter else 'All Cycles'
    }

    # Get all strategic cycles for dropdown
    all_cycles = qs.values_list(
        'strategic_report_period__action_plan__strategic_cycle__name', flat=True
    ).distinct()

    return render(request, 'swot_report/chart.html', {
        'plot_html_swot_type': fig_swot_type.to_html(full_html=False),
        'plot_html_priority': fig_priority.to_html(full_html=False),
        'plot_html_impact': fig_impact.to_html(full_html=False),
        'plot_html_likelihood': fig_likelihood.to_html(full_html=False),
        'plot_html_pillar': fig_pillar.to_html(full_html=False),
        'summary_data': summary_data,
        'all_cycles': all_cycles,
        'selected_cycle': cycle_filter
    })
